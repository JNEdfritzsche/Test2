import streamlit as st
import pandas as pd
from collections import defaultdict, deque
from io import BytesIO
import hashlib
import urllib.parse

from streamlit_vis_network import streamlit_vis_network

#To run locally: python -m streamlit run .\streamlit_app.py

# ============================================================
# Core Routing (Version A logic, refactored to use DataFrames)
# ============================================================

class CableNetwork:
    def __init__(self):
        self.graph = defaultdict(list)
        self.noise_levels = {}
        self.endpoints = defaultdict(list)
        self.endpoint_exposed = {}
        self.edge_exposed = {}  # (sorted_a, sorted_b) -> bool
        self.edge_exposed = {}  # (sorted_a, sorted_b) -> bool

    @staticmethod
    def _infer_noise_from_name(run_name: str):
        name = str(run_name or "")
        if "T6" in name:
            return {1}
        if "T4" in name:
            return {2}
        if "T3" in name:
            return {3}
        return set()

    @staticmethod
    def _parse_noise_levels(val, run_name=None):
        if pd.isna(val) or str(val).strip() == "":
            return CableNetwork._infer_noise_from_name(run_name)

        if isinstance(val, (int, float)) and not pd.isna(val):
            return {int(val)}

        s = str(val).strip()
        parts = [p.strip() for p in s.split(",")]
        out = set()
        for p in parts:
            if p:
                try:
                    out.add(int(float(p)))
                except ValueError:
                    pass
        return out

    def add_tray_or_conduit(self, name, noise_levels):
        levels = set(noise_levels) if isinstance(noise_levels, (set, list, tuple)) else {int(noise_levels)}
        self.noise_levels[name] = levels
        if name not in self.graph:
            self.graph[name] = []

    def connect(self, from_node, to_node, exposed=False):
        if from_node in self.noise_levels and to_node in self.noise_levels:
            self.graph[from_node].append((to_node, self.noise_levels[to_node]))
            self.graph[to_node].append((from_node, self.noise_levels[from_node]))
            edge_key = tuple(sorted([from_node, to_node]))
            if exposed:
                self.edge_exposed[edge_key] = True

    def register_endpoint(self, device_name, tray_or_conduits, exposed=None):
        device_name = str(device_name).strip().lstrip("+")
        for item in str(tray_or_conduits).split(","):
            it = item.strip()
            if it:
                self.endpoints[device_name].append(it)
        if exposed is not None:
            self.endpoint_exposed[device_name] = bool(exposed)

    def find_route(self, starts, ends, allowed_noise_level, include_nodes=None, exclude_nodes=None):
        """
        Find a route from starts to ends with allowed noise level.
        
        Args:
            starts: List of starting nodes
            ends: Set of ending nodes
            allowed_noise_level: Required noise level for the path
            include_nodes: If provided, route must pass through ALL of these nodes
            exclude_nodes: If provided, route must avoid these nodes
        
        Returns:
            List representing the path, or None if no valid route exists
        """
        # Normalize include/exclude sets
        if include_nodes is None:
            include_nodes = set()
        else:
            include_nodes = set(n.strip() for n in include_nodes if n.strip())
        
        if exclude_nodes is None:
            exclude_nodes = set()
        else:
            exclude_nodes = set(n.strip() for n in exclude_nodes if n.strip())
        
        end_set = set(ends)
        
        # If no include_nodes specified, use simple BFS
        if not include_nodes:
            visited = set()
            queue = deque([[start] for start in starts])
            
            while queue:
                path = queue.popleft()
                node = path[-1]
                if node in end_set:
                    return path
                if node not in visited:
                    visited.add(node)
                    for neighbor, neighbor_levels in self.graph[node]:
                        if neighbor not in visited and allowed_noise_level in (neighbor_levels or set()):
                            if neighbor in exclude_nodes:
                                continue
                            queue.append(path + [neighbor])
            return None
        
        # With include_nodes: find path through ALL include nodes
        # Strategy: route from start -> each include node -> end, ensuring we visit all
        include_list = list(include_nodes)
        
        # Helper: find route between two node sets
        def find_path_between(starts_set, ends_set, exclude_set, visited_global=None):
            if visited_global is None:
                visited_global = set()
            visited = set()
            queue = deque([[start] for start in starts_set])
            
            while queue:
                path = queue.popleft()
                node = path[-1]
                if node in ends_set:
                    return path
                if node not in visited:
                    visited.add(node)
                    for neighbor, neighbor_levels in self.graph[node]:
                        if neighbor not in visited and allowed_noise_level in (neighbor_levels or set()):
                            if neighbor in exclude_set:
                                continue
                            queue.append(path + [neighbor])
            return None
        
        # Try to build a path that visits all include_nodes
        # Start from start_trays, go through all include_nodes, then to end_trays
        current_starts = set(starts)
        full_path = []
        remaining_includes = set(include_list)
        
        while remaining_includes:
            # Find closest include_node from current position
            closest_path = None
            closest_node = None
            
            for inc_node in remaining_includes:
                path = find_path_between(current_starts, {inc_node}, exclude_nodes)
                if path:
                    if closest_path is None or len(path) < len(closest_path):
                        closest_path = path
                        closest_node = inc_node
            
            if closest_path is None:
                # Can't reach any remaining include node
                return None
            
            # Add path to full path (avoiding duplication at junction)
            if full_path:
                full_path.extend(closest_path[1:])
            else:
                full_path = closest_path
            
            current_starts = {closest_node}
            remaining_includes.discard(closest_node)
        
        # Now route from last include node to end
        final_path = find_path_between(current_starts, end_set, exclude_nodes)
        if final_path is None:
            return None
        
        # Combine: full_path already ends at the last include node
        if len(final_path) > 1:
            full_path.extend(final_path[1:])
        
        return full_path


    def build_from_dfs(self, trays_df: pd.DataFrame, connections_df: pd.DataFrame, endpoints_df: pd.DataFrame):
        for _, row in trays_df.iterrows():
            run_name = row.get("RunName", None)
            if pd.isna(run_name) or str(run_name).strip() == "":
                continue
            run_name = str(run_name).strip()
            levels = self._parse_noise_levels(row.get("Noise Level", None), run_name=run_name)
            if levels:
                self.add_tray_or_conduit(run_name, levels)

        for _, row in connections_df.iterrows():
            a = row.get("From", None)
            b = row.get("To", None)
            if pd.isna(a) or pd.isna(b):
                continue
            exposed_flag = False
            try:
                val = row.get("Exposed?", "")
                if str(val).strip().lower() == "yes":
                    exposed_flag = True
            except Exception:
                pass
            self.connect(str(a).strip(), str(b).strip(), exposed=exposed_flag)

        for _, row in endpoints_df.iterrows():
            device = row.get("device/panel", "")
            trays = row.get("tray/conduit(s)", "")
            exposed_flag = False
            try:
                val = row.iloc[2] if len(row) >= 3 else ""
            except Exception:
                val = ""
            if str(val).strip().lower() == "yes":
                exposed_flag = True
            self.register_endpoint(device, trays, exposed=exposed_flag)

    def route_cables_df(self, cables_df: pd.DataFrame) -> pd.DataFrame:
        cables = []
        seen = set()

        for _, row in cables_df.iterrows():
            sort = None
            if "Sort" in row and not pd.isna(row["Sort"]):
                try:
                    sort = int(row["Sort"])
                except Exception:
                    sort = None

            name = str(row.get("Cable number", ""))
            start = str(row.get("equipfrom", "")).strip().lstrip("+")
            end = str(row.get("equipto", "")).strip().lstrip("+")

            nl = row.get("Noise Level", None)
            if pd.isna(nl):
                continue
            try:
                noise_level = int(float(nl))
            except Exception:
                continue

            # Extract INCLUDE and EXCLUDE columns (handle NaN values properly)
            include_val = row.get("INCLUDE", "")
            exclude_val = row.get("EXCLUDE", "")
            
            include_str = "" if pd.isna(include_val) else str(include_val).strip()
            exclude_str = "" if pd.isna(exclude_val) else str(exclude_val).strip()
            
            include_nodes = [n.strip() for n in include_str.split(",") if n.strip()] if include_str else []
            exclude_nodes = [n.strip() for n in exclude_str.split(",") if n.strip()] if exclude_str else []

            key = sort if sort is not None else name
            if key in seen:
                continue
            seen.add(key)

            cables.append({
                "sort": sort,
                "name": name,
                "start": start,
                "end": end,
                "noise_level": noise_level,
                "include_nodes": include_nodes,
                "exclude_nodes": exclude_nodes
            })

        results = []
        for cable in cables:
            nl = cable["noise_level"]
            include_nodes = cable["include_nodes"]
            exclude_nodes = cable["exclude_nodes"]
            warnings = []

            # Special case: route is internal if source and destination are the same
            if cable["start"] == cable["end"]:
                path_result = "INTERNAL"
            else:
                start_trays = [t for t in self.endpoints.get(cable["start"], [])
                               if nl in self.noise_levels.get(t, set())]
                end_trays = [t for t in self.endpoints.get(cable["end"], [])
                             if nl in self.noise_levels.get(t, set())]

                if not start_trays and not end_trays:
                    path_result = "Error: No endpoints with matching noise level (start & end)"
                elif not start_trays:
                    path_result = "Error: No start endpoint with matching noise level"
                elif not end_trays:
                    path_result = "Error: No end endpoint with matching noise level"
                else:
                    # Validate INCLUDE nodes have matching noise level
                    if include_nodes:
                        for node in include_nodes:
                            node_levels = self.noise_levels.get(node, set())
                            if nl not in node_levels:
                                warnings.append(f"INCLUDE node '{node}' does not support noise level {nl} (supports: {node_levels if node_levels else 'none'})")
                    
                    # Try to find route with constraints
                    path = self.find_route(start_trays, set(end_trays), nl, 
                                          include_nodes=include_nodes if include_nodes else None,
                                          exclude_nodes=exclude_nodes if exclude_nodes else None)
                    
                    if path:
                        suffix = ""
                        if nl == 1:
                            suffix = "(T6)"
                        elif nl == 2:
                            suffix = "(T4)"

                        def format_node(node):
                            node = str(node)
                            if "LT" in node:
                                return f"{node}{suffix}"
                            if "CND" in node:
                                return node
                            return node

                        parts = [format_node(p) for p in path]

                        FROM_exposed = self.endpoint_exposed.get(cable["start"], False)
                        TO_exposed = self.endpoint_exposed.get(cable["end"], False)
                        
                        # Check for exposed edges along the path
                        for i in range(len(path) - 1):
                            edge_key = tuple(sorted([path[i], path[i+1]]))
                            if self.edge_exposed.get(edge_key, False):
                                parts[i+1] = "EXPOSED CONDUIT ROUTE," + parts[i+1]

                        if FROM_exposed:
                            parts.insert(0, "EXPOSED CONDUIT ROUTE")
                        if TO_exposed:
                            parts.append("EXPOSED CONDUIT ROUTE")

                        path_result = ",".join(parts)
                    else:
                        # No route found; provide diagnostic message
                        if include_nodes:
                            path_result = f"Error: No valid route using only INCLUDE nodes: {', '.join(include_nodes)}"
                        elif exclude_nodes:
                            path_result = f"Error: No valid route while avoiding EXCLUDE nodes: {', '.join(exclude_nodes)}"
                        else:
                            path_result = "No valid route"


            result_entry = {
                "Sort": cable["sort"],
                "Cable number": cable["name"],
                "equipfrom": cable["start"],
                "equipto": cable["end"],
                "Noise Level": nl,
                "Via": path_result
            }
            
            # Append warnings as a separate column if any exist
            if warnings:
                result_entry["Warnings"] = "; ".join(warnings)
            
            results.append(result_entry)

        return pd.DataFrame(results)


# ============================================================
# Reverse Engineering (from routed workbook)
# ============================================================

import re
from itertools import pairwise

SUFFIX_RE = re.compile(r"\s*\((?:T6|T4)\)$")
EXPOSED_SENTINEL = "EXPOSED CONDUIT ROUTE"

def normalize_node(name: str) -> str:
    """Strip display suffixes '(T6)/(T4)' from Via tokens to get the canonical node id."""
    return SUFFIX_RE.sub("", str(name or "").strip())

def levels_from_suffix(token: str) -> set:
    """Return {1} if '(T6)' present, {2} if '(T4)' present, else empty set."""
    s = str(token or "")
    levels = set()
    if "(T6)" in s:
        levels.add(1)
    if "(T4)" in s:
        levels.add(2)
    return levels

def clamp_valid_levels(levels: set) -> set:
    """Keep only valid levels in {1,2,3,4}."""
    return {lvl for lvl in (levels or set()) if lvl in {1,2,3,4}}

def levels_to_cell_value(levels: set):
    """
    Convert a set to what the Tray sheet expects:
    - {1} -> "1"
    - {1,2} -> "1,2"
    - empty -> ""
    """
    if not levels:
        return ""
    return ",".join(str(x) for x in sorted(levels))

def reverse_engineer_from_routes(xlsx_path: str):
    """
    Reads 'CableRoutes(output)' sheet and reconstructs:
      - Tray sheet: RunName, Noise Level (supports multi-level like "1,2")
      - Connections sheet: From, To
      - Endpoints sheet:
          * One row per device/panel
          * tray/conduit(s): ALL possible endpoint trays for that device, comma-separated, NO SPACES
          * Exposed?: "yes" if any occurrence for that device had EXPOSED at that end

    Also returns the original Cables(input) + CableRoutes(output) DataFrames (if present)
    so they can be written into the reconstructed workbook unchanged.
    """
    xls = pd.ExcelFile(xlsx_path)

    # --- Original sheets to carry through (if present) ---
    if "CableRoutes(output)" not in xls.sheet_names:
        raise ValueError("Uploaded workbook is missing 'CableRoutes(output)' sheet.")

    df_routes_original = pd.read_excel(xls, sheet_name="CableRoutes(output)")
    df_routes = df_routes_original

    df_cables_input_original = None
    if "Cables(input)" in xls.sheet_names:
        df_cables_input_original = pd.read_excel(xls, sheet_name="Cables(input)")

    trays_seen = set()
    connections = set()  # undirected (a,b) with a<b
    exposed_edges = set()  # edges marked as exposed (a,b) with a<b
    node_levels = {}     # node -> set of supported levels

    # device -> set of endpoint trays
    device_to_trays = {}
    # device -> exposed_yes_bool (aggregated)
    device_exposed_yes = {}

    for _, row in df_routes.iterrows():
        via_raw = str(row.get("Via", "")).strip()
        if not via_raw or "Error" in via_raw or "No valid route" in via_raw:
            continue

        # Column E: Noise Level (per cable)
        cable_levels = set()
        try:
            nl = int(row.get("Noise Level", None))
            if nl in {1, 2, 3, 4}:
                cable_levels.add(nl)
        except Exception:
            pass

        # Parse route from Via
        tokens_raw_all = [p.strip() for p in via_raw.split(",") if p.strip()]
        if not tokens_raw_all:
            continue

        tokens_norm_all = [normalize_node(p) for p in tokens_raw_all]
        is_exposed_all = [t.strip().upper() == EXPOSED_SENTINEL for t in tokens_norm_all]

        # Filter out EXPOSED entirely for trays, node_levels, and connections
        filtered_pairs = [(r, n) for r, n, ex in zip(tokens_raw_all, tokens_norm_all, is_exposed_all) if not ex]
        tokens_raw = [r for r, _ in filtered_pairs]
        tokens_norm = [n for _, n in filtered_pairs]

        # If filtering removed everything, we cannot build endpoints or network
        if not tokens_norm:
            continue

        # Track nodes + noise evidence (only for non-EXPOSED nodes)
        for raw_tok, norm_tok in zip(tokens_raw, tokens_norm):
            trays_seen.add(norm_tok)
            inferred = levels_from_suffix(raw_tok) | cable_levels
            if inferred:
                node_levels.setdefault(norm_tok, set()).update(inferred)

        # Build undirected edges among non-EXPOSED nodes
        # Also detect exposed edges: if an EXPOSED token appears between two nodes, mark that edge as exposed
        for i, (a, b) in enumerate(pairwise(tokens_norm)):
            edge = tuple(sorted([a, b]))
            connections.add(edge)
            
            # Check if there was an EXPOSED token between these nodes in the original sequence
            # We need to check the original is_exposed_all to see if any EXPOSED came between them
            # Map indices: if we're at pair (tokens_norm[i], tokens_norm[i+1]), 
            # we need to check all original indices to find the indices of these tokens
            # and see if any EXPOSED was between them
            try:
                idx_a = tokens_norm_all.index(a)
                idx_b = tokens_norm_all.index(b)
                # Check if any EXPOSED is between these indices in original
                has_exposed_between = any(is_exposed_all[j] for j in range(min(idx_a, idx_b) + 1, max(idx_a, idx_b)))
                if has_exposed_between:
                    exposed_edges.add(edge)
            except Exception:
                pass

        # --- Endpoint assignment with EXPOSED awareness for "use" flag ---
        start_dev = str(row.get("equipfrom", "")).strip().lstrip("+")
        end_dev   = str(row.get("equipto", "")).strip().lstrip("+")

        start_had_exposed = (len(is_exposed_all) > 0 and is_exposed_all[0] is True)
        end_had_exposed   = (len(is_exposed_all) > 0 and is_exposed_all[-1] is True)

        start_endpoint = tokens_norm[0]
        end_endpoint   = tokens_norm[-1]

        if start_dev:
            device_to_trays.setdefault(start_dev, set()).add(start_endpoint)
            device_exposed_yes[start_dev] = bool(device_exposed_yes.get(start_dev, False) or start_had_exposed)

        if end_dev:
            device_to_trays.setdefault(end_dev, set()).add(end_endpoint)
            device_exposed_yes[end_dev] = bool(device_exposed_yes.get(end_dev, False) or end_had_exposed)

    # Clamp to valid levels and fill missing nodes with empty sets
    for n in trays_seen:
        node_levels[n] = clamp_valid_levels(node_levels.get(n, set()))

    # Build DataFrames
    trays_df = pd.DataFrame(
        [{"RunName": name, "Noise Level": levels_to_cell_value(node_levels.get(name, set()))}
         for name in sorted(trays_seen)]
    )
    connections_df = pd.DataFrame(
        [{"From": a, "To": b, "Exposed?": "yes" if (a, b) in exposed_edges else ""}
         for a, b in sorted(connections)]
    )

    # Endpoints: one row per device, trays comma-separated with NO spaces
    endpoint_rows = []
    for dev in sorted(device_to_trays.keys()):
        trays_list = sorted(device_to_trays[dev])
        trays_joined = ",".join(trays_list)  # IMPORTANT: no spaces
        endpoint_rows.append({
            "device/panel": dev,
            "tray/conduit(s)": trays_joined,
            "Exposed?": "yes" if device_exposed_yes.get(dev, False) else "no"
        })

    endpoints_df = pd.DataFrame(endpoint_rows)

    return (
        trays_df,
        connections_df,
        endpoints_df,
        node_levels,
        connections,
        df_cables_input_original,
        df_routes_original,
    )

# ============================================================
# Excel I/O helpers
# ============================================================

REQUIRED_SHEETS = ["Tray", "Connections", "Endpoints", "Cables(input)"]

def load_excel_to_dfs(file_bytes: bytes) -> tuple[dict, bool]:
    """
    Load Excel file into dataframes.
    Returns: (dfs dict, is_older_template bool)
    is_older_template is True if INCLUDE/EXCLUDE columns were added to Cables(input)
    """
    xls = pd.ExcelFile(BytesIO(file_bytes))
    dfs = {}
    is_older_template = False
    
    for sh in REQUIRED_SHEETS:
        if sh not in xls.sheet_names:
            raise ValueError(f"Missing required sheet: {sh}")
        converters = None
        if sh == "Tray":
            converters = {"RunName": str, "Noise Level": str}
        elif sh == "Connections":
            converters = {"From": str, "To": str, "Exposed?": str}
        elif sh == "Endpoints":
            converters = {"device/panel": str, "tray/conduit(s)": str, "Exposed?": str}
        elif sh == "Cables(input)":
            converters = {
                "Cable number": str,
                "equipfrom": str,
                "equipto": str,
                "Noise Level": str,
                "Sort": str,
            }

        dfs[sh] = pd.read_excel(xls, sh, converters=converters)
    
    if "Tray" in dfs:
        dfs["Tray"] = ensure_person_columns(ensure_xy_columns(dfs["Tray"]))
    
    # Ensure Connections has "Exposed?" column (for backwards compatibility)
    if "Connections" in dfs:
        if "Exposed?" not in dfs["Connections"].columns and "Exposed Conduit Route?" not in dfs["Connections"].columns:
            dfs["Connections"]["Exposed?"] = ""
        elif "Exposed Conduit Route?" in dfs["Connections"].columns and "Exposed?" not in dfs["Connections"].columns:
            dfs["Connections"].rename(columns={"Exposed Conduit Route?": "Exposed?"}, inplace=True)
    
    # Ensure Cables(input) has INCLUDE and EXCLUDE columns (for backwards compatibility)
    if "Cables(input)" in dfs:
        if "INCLUDE" not in dfs["Cables(input)"].columns:
            dfs["Cables(input)"]["INCLUDE"] = ""
            is_older_template = True
        if "EXCLUDE" not in dfs["Cables(input)"].columns:
            dfs["Cables(input)"]["EXCLUDE"] = ""
            is_older_template = True
    
    return dfs, is_older_template

def write_updated_workbook_bytes(dfs: dict) -> bytes:
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sh, df in dfs.items():
            df.to_excel(writer, sheet_name=sh, index=False)
    out.seek(0)
    return out.getvalue()

def write_routed_workbook_bytes(dfs: dict, routes_df: pd.DataFrame) -> bytes:
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sh, df in dfs.items():
            df.to_excel(writer, sheet_name=sh, index=False)

        out_df = routes_df.copy()
        if "Cable number" in out_df.columns:
            out_df = out_df.rename(columns={"Cable number": "Cable Number"})
        out_df = out_df[["Sort", "Cable Number", "equipfrom", "equipto", "Noise Level", "Via"]]
        out_df = out_df.where(out_df.notna(), "")
        for c in out_df.columns:
            out_df[c] = out_df[c].astype(str)

        out_df.to_excel(writer, sheet_name="CableRoutes(output)", index=False)

        from openpyxl.styles import numbers
        wb = writer.book
        ws = wb["CableRoutes(output)"]
        header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        for col_name in ["Cable Number", "equipfrom", "equipto", "Via"]:
            if col_name in header_map:
                cidx = header_map[col_name]
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=cidx)
                    cell.value = "" if cell.value is None else str(cell.value)
                    cell.number_format = numbers.FORMAT_TEXT

    out.seek(0)
    return out.getvalue()

def validate_dfs(dfs: dict) -> list[str]:
    errors = []
    expected_cols = {
        "Tray": {"RunName"},
        "Connections": {"From", "To"},
        "Endpoints": {"device/panel", "tray/conduit(s)"},
        "Cables(input)": {"Cable number", "equipfrom", "equipto", "Noise Level"},
    }
    
    # Ensure Connections has "Exposed?" column (add if missing)
    if "Connections" in dfs:
        if "Exposed?" not in dfs["Connections"].columns and "Exposed Conduit Route?" not in dfs["Connections"].columns:
            dfs["Connections"]["Exposed?"] = ""
        elif "Exposed Conduit Route?" in dfs["Connections"].columns and "Exposed?" not in dfs["Connections"].columns:
            dfs["Connections"].rename(columns={"Exposed Conduit Route?": "Exposed?"}, inplace=True)
    
    for sh, cols in expected_cols.items():
        missing = cols - set(dfs[sh].columns)
        if missing:
            errors.append(f"{sh}: missing columns {sorted(missing)}")

    try:
        tray_df = dfs.get("Tray")
        con_df = dfs.get("Connections")
        if tray_df is not None and con_df is not None and not tray_df.empty and not con_df.empty:
            tray_names = set(tray_df["RunName"].astype(str).str.strip())
            from_set = set(con_df["From"].astype(str).str.strip())
            to_set = set(con_df["To"].astype(str).str.strip())
            missing_nodes = sorted(n for n in (from_set | to_set) if n and n not in tray_names)
            if missing_nodes:
                errors.append("Connections: missing Tray.RunName entries for " + ", ".join(missing_nodes))
    except Exception:
        pass

    return errors

def build_demo_workbook_bytes() -> bytes:
    tray = pd.DataFrame([
        {"RunName": "Alex", "Noise Level": "1",   "Email": "alex@example.com",   "Phone": "555-0101", "Role": "Manager",   "Notes": "Team lead",         "X": pd.NA, "Y": pd.NA},
        {"RunName": "Blair", "Noise Level": "1",   "Email": "blair@example.com",  "Phone": "555-0102", "Role": "Developer", "Notes": "Frontend focus",    "X": pd.NA, "Y": pd.NA},
        {"RunName": "Casey", "Noise Level": "1",   "Email": "casey@example.com",  "Phone": "555-0103", "Role": "Analyst",   "Notes": "Reports and data",  "X": pd.NA, "Y": pd.NA},
        {"RunName": "Drew","Noise Level": "1",   "Email": "drew@example.com",   "Phone": "555-0104", "Role": "Designer",  "Notes": "Visual systems",   "X": pd.NA, "Y": pd.NA},

        {"RunName": "Erin", "Noise Level": "2",   "Email": "erin@example.com",   "Phone": "555-0105", "Role": "Developer", "Notes": "Backend services", "X": pd.NA, "Y": pd.NA},
        {"RunName": "Flynn", "Noise Level": "2",   "Email": "flynn@example.com",  "Phone": "555-0106", "Role": "Sales",     "Notes": "Client contact",   "X": pd.NA, "Y": pd.NA},
        {"RunName": "Gray", "Noise Level": "2",   "Email": "gray@example.com",   "Phone": "555-0107", "Role": "HR",        "Notes": "People ops",       "X": pd.NA, "Y": pd.NA},
        {"RunName": "Harper","Noise Level": "2",   "Email": "harper@example.com", "Phone": "555-0108", "Role": "Designer",  "Notes": "Brand work",       "X": pd.NA, "Y": pd.NA},

        {"RunName": "Indy", "Noise Level": "1,2", "Email": "indy@example.com",   "Phone": "555-0109", "Role": "Manager",   "Notes": "Cross-team owner", "X": pd.NA, "Y": pd.NA},
        {"RunName": "Jules","Noise Level": "1,2", "Email": "jules@example.com",  "Phone": "555-0110", "Role": "Analyst",   "Notes": "Operations",       "X": pd.NA, "Y": pd.NA},

        {"RunName": "Kai", "Noise Level": "1",   "Email": "kai@example.com",    "Phone": "555-0111", "Role": "Developer", "Notes": "Automation",       "X": pd.NA, "Y": pd.NA},
        {"RunName": "Lane", "Noise Level": "2",   "Email": "lane@example.com",   "Phone": "555-0112", "Role": "Other",     "Notes": "Contractor",       "X": pd.NA, "Y": pd.NA},
    ])

    connections = pd.DataFrame([
        {"From": "Alex", "To": "Blair", "Exposed?": ""},
        {"From": "Blair", "To": "Casey", "Exposed?": ""},
        {"From": "Blair", "To": "Kai", "Exposed?": ""},
        {"From": "Casey", "To": "Drew", "Exposed?": ""},

        {"From": "Erin", "To": "Flynn", "Exposed?": ""},
        {"From": "Flynn", "To": "Gray", "Exposed?": ""},
        {"From": "Flynn", "To": "Lane", "Exposed?": ""},
        {"From": "Gray", "To": "Harper", "Exposed?": ""},

        {"From": "Drew", "To": "Indy", "Exposed?": ""},
        {"From": "Harper", "To": "Indy", "Exposed?": ""},
        {"From": "Indy",  "To": "Jules", "Exposed?": ""},

        {"From": "Kai",  "To": "Indy", "Exposed?": ""},
        {"From": "Lane",  "To": "Indy", "Exposed?": ""},
    ])

    endpoints = pd.DataFrame([
        {"device/panel": "PANEL-A", "tray/conduit(s)": "Alex,Erin", "Exposed?": ""},
        {"device/panel": "PANEL-B", "tray/conduit(s)": "Casey,Gray", "Exposed?": "yes"},
        {"device/panel": "PANEL-C", "tray/conduit(s)": "Kai",       "Exposed?": ""},
        {"device/panel": "PANEL-D", "tray/conduit(s)": "Lane",       "Exposed?": ""},
    ])

    cables = pd.DataFrame([
        {"Sort": 1, "Cable number": "+CBL-0001", "equipfrom": "PANEL-A", "equipto": "PANEL-B", "Noise Level": 1, "INCLUDE": "", "EXCLUDE": ""},
        {"Sort": 2, "Cable number": "+CBL-0002", "equipfrom": "PANEL-A", "equipto": "PANEL-B", "Noise Level": 2, "INCLUDE": "", "EXCLUDE": ""},
        {"Sort": 3, "Cable number": "+CBL-0003", "equipfrom": "PANEL-C", "equipto": "PANEL-B", "Noise Level": 1, "INCLUDE": "", "EXCLUDE": ""},
        {"Sort": 4, "Cable number": "+CBL-0004", "equipfrom": "PANEL-D", "equipto": "PANEL-B", "Noise Level": 2, "INCLUDE": "", "EXCLUDE": ""},
    ])

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        tray.to_excel(writer, sheet_name="Tray", index=False)
        connections.to_excel(writer, sheet_name="Connections", index=False)
        endpoints.to_excel(writer, sheet_name="Endpoints", index=False)
        cables.to_excel(writer, sheet_name="Cables(input)", index=False)
    out.seek(0)
    return out.getvalue()


# ============================================================
# Endpoint lookup helpers
# ============================================================

def _normalize_token_list(s: str) -> list[str]:
    parts = [p.strip() for p in str(s or "").split(",")]
    return [p for p in parts if p]

def build_endpoint_to_trays_map(endpoints_df: pd.DataFrame) -> dict[str, list[str]]:
    m: dict[str, list[str]] = {}
    if endpoints_df is None or endpoints_df.empty:
        return m
    if "device/panel" not in endpoints_df.columns or "tray/conduit(s)" not in endpoints_df.columns:
        return m

    for _, r in endpoints_df.iterrows():
        ep = str(r.get("device/panel", "")).strip().lstrip("+")
        tc = r.get("tray/conduit(s)", "")
        if not ep:
            continue
        trays = _normalize_token_list(tc)
        if ep in m:
            m[ep].extend(trays)
        else:
            m[ep] = trays

    for k, v in list(m.items()):
        seen = set()
        out = []
        for x in v:
            if x not in seen:
                out.append(x)
                seen.add(x)
        m[k] = out
    return m

def build_tray_to_endpoints_map(endpoints_df: pd.DataFrame) -> dict[str, list[str]]:
    m: dict[str, list[str]] = {}
    if endpoints_df is None or endpoints_df.empty:
        return m
    if "device/panel" not in endpoints_df.columns or "tray/conduit(s)" not in endpoints_df.columns:
        return m

    for _, r in endpoints_df.iterrows():
        ep = str(r.get("device/panel", "")).strip().lstrip("+")
        tc = r.get("tray/conduit(s)", "")
        if not ep:
            continue
        trays = _normalize_token_list(tc)
        for t in trays:
            m.setdefault(t, []).append(ep)

    for k, v in list(m.items()):
        seen = set()
        out = []
        for x in v:
            if x not in seen:
                out.append(x)
                seen.add(x)
        m[k] = out
    return m


# ============================================================
# Route lookup helpers
# ============================================================

def compute_routes_df(tray_df: pd.DataFrame, connections_df: pd.DataFrame, endpoints_df: pd.DataFrame, cables_df: pd.DataFrame) -> pd.DataFrame:
    net = CableNetwork()
    net.build_from_dfs(tray_df, connections_df, endpoints_df)
    return net.route_cables_df(cables_df)

def _strip_route_suffix(node: str) -> str:
    s = str(node or "").strip()
    for suf in ["(T6)", "(T4)", "(T3)"]:
        if s.endswith(suf):
            s = s[: -len(suf)].strip()
    return s

def nodes_from_via_string(via: str) -> list[str]:
    if via is None or str(via).strip() == "":
        return []
    s = str(via).strip()
    if s.lower().startswith("error:"):
        return []
    if s.strip().lower() == "no valid route":
        return []
    parts = [p.strip() for p in s.split(",") if p.strip()]
    cleaned = []
    for p in parts:
        if p.upper() == "EXPOSED CONDUIT ROUTE":
            continue
        cleaned.append(_strip_route_suffix(p))
    return cleaned

def format_route_nodes_for_via(nodes: list[str], noise_level: int | None = None) -> str:
    if not nodes:
        return ""
    suffix = ""
    if noise_level == 1:
        suffix = "(T6)"
    elif noise_level == 2:
        suffix = "(T4)"

    def format_node(node: str) -> str:
        node = str(node)
        if "LT" in node and suffix:
            return f"{node}{suffix}"
        if "CND" in node:
            return node
        return node

    return ",".join(format_node(n) for n in nodes)


# ============================================================
# Graph helpers (SVG nodes + seam fix + highlighting)
# ============================================================

ORANGE = "#FFA500"
GREEN  = "#00A651"
YELLOW = "#FFD200"
GRAY   = "#CFCFCF"

EDGE_COLOR = "#000000"
EDGE_WIDTH = 2

NODE_BORDER_COLOR = "#333333"
SVG_BORDER_WIDTH = 2

TRAY_SIDE = 70
TRAY_RADIUS = 14
IMAGE_SIZE = 32

PROBE_ID = "__POS_PROBE__"

PERSON_COLUMNS = ["Email", "Phone", "Role", "Notes"]

HIGHLIGHT_BORDER_WIDTH = 10
HIGHLIGHT_BORDER_COLOR = "#FF00FF"
HIGHLIGHT_GLOW_SIZE = 22

def ensure_xy_columns(tray_df: pd.DataFrame) -> pd.DataFrame:
    tray_df = tray_df.copy()
    if "X" not in tray_df.columns:
        tray_df["X"] = pd.NA
    if "Y" not in tray_df.columns:
        tray_df["Y"] = pd.NA
    return tray_df

def ensure_person_columns(tray_df: pd.DataFrame) -> pd.DataFrame:
    tray_df = tray_df.copy()
    for col in PERSON_COLUMNS:
        if col not in tray_df.columns:
            tray_df[col] = ""
    return tray_df

def tray_has_any_xy(tray_df: pd.DataFrame) -> bool:
    if tray_df is None or tray_df.empty:
        return False
    df = ensure_xy_columns(tray_df)
    xs = pd.to_numeric(df["X"], errors="coerce")
    ys = pd.to_numeric(df["Y"], errors="coerce")
    mask = xs.notna() & ys.notna()
    return bool(mask.any())

def noise_title(run_name: str, noise_val) -> str:
    lv = CableNetwork._parse_noise_levels(noise_val, run_name=run_name)
    if not lv:
        return "Noise Levels: N/A"
    return "Noise Levels: " + ",".join(str(x) for x in sorted(lv))

def infer_type_from_name(name: str) -> str:
    s = str(name).upper()
    if "EP" in s:
        return "Manhole"
    if "CND" in s:
        return "Conduit"
    if "LT" in s:
        return "Tray"
    return "Node"

def svg_data_uri(svg: str) -> str:
    return "data:image/svg+xml;utf8," + urllib.parse.quote(svg)

def person_circle_svg(diameter: int, color: str, border: str = NODE_BORDER_COLOR) -> str:
    r = diameter / 2
    cx = diameter / 2
    cy = diameter / 2
    head_r = diameter * 0.16
    shoulder_w = diameter * 0.48
    shoulder_h = diameter * 0.24
    shoulder_x = (diameter - shoulder_w) / 2
    shoulder_y = diameter * 0.54
    return f"""
    <svg xmlns="http://www.w3.org/2000/svg" width="{diameter}" height="{diameter}" viewBox="0 0 {diameter} {diameter}">
      <circle cx="{cx}" cy="{cy}" r="{r - 1}" fill="{color}"/>
      <circle cx="{cx}" cy="{cy}" r="{r - 1}" fill="none" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
      <circle cx="{cx}" cy="{diameter * 0.34}" r="{head_r}" fill="rgba(255,255,255,0.92)"/>
      <rect x="{shoulder_x}" y="{shoulder_y}" width="{shoulder_w}" height="{shoulder_h}" rx="{shoulder_h / 2}" fill="rgba(255,255,255,0.92)"/>
    </svg>
    """.strip()

def solid_rounded_square_svg(side: int, r: int, color: str, border: str = NODE_BORDER_COLOR) -> str:
    return f"""
    <svg xmlns="http://www.w3.org/2000/svg" width="{side}" height="{side}" viewBox="0 0 {side} {side}">
      <rect x="0" y="0" width="{side}" height="{side}" rx="{r}" ry="{r}" fill="{color}"/>
      <rect x="1" y="1" width="{side-2}" height="{side-2}" rx="{r}" ry="{r}"
            fill="none" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
    </svg>
    """.strip()

def split_rounded_square_svg(side: int, r: int, left_color: str, right_color: str, border: str = NODE_BORDER_COLOR) -> str:
    half = side // 2
    left_w = half + 1
    right_x = half
    right_w = side - half
    return f"""
    <svg xmlns="http://www.w3.org/2000/svg" width="{side}" height="{side}" viewBox="0 0 {side} {side}">
      <defs>
        <clipPath id="clipR">
          <rect x="0" y="0" width="{side}" height="{side}" rx="{r}" ry="{r}" />
        </clipPath>
      </defs>
      <g clip-path="url(#clipR)">
        <rect x="0" y="0" width="{left_w}" height="{side}" fill="{left_color}"/>
        <rect x="{right_x}" y="0" width="{right_w}" height="{side}" fill="{right_color}"/>
      </g>
      <rect x="1" y="1" width="{side-2}" height="{side-2}" rx="{r}" ry="{r}"
            fill="none" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
    </svg>
    """.strip()

def solid_circle_svg(d: int, color: str, border: str = NODE_BORDER_COLOR) -> str:
    r = d / 2
    rr = r - 1
    return f"""
    <svg xmlns="http://www.w3.org/2000/svg" width="{d}" height="{d}" viewBox="0 0 {d} {d}">
      <circle cx="{r}" cy="{r}" r="{rr}" fill="{color}"/>
      <circle cx="{r}" cy="{r}" r="{rr}" fill="none" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
    </svg>
    """.strip()

def split_circle_svg(d: int, left_color: str, right_color: str, border: str = NODE_BORDER_COLOR) -> str:
    r = d / 2
    rr = r - 1
    left_w = int(r) + 1
    right_x = int(r)
    right_w = d - int(r)
    return f"""
    <svg xmlns="http://www.w3.org/2000/svg" width="{d}" height="{d}" viewBox="0 0 {d} {d}">
      <defs>
        <clipPath id="clipC">
          <circle cx="{r}" cy="{r}" r="{rr}"/>
        </clipPath>
      </defs>
      <g clip-path="url(#clipC)">
        <rect x="0" y="0" width="{left_w}" height="{d}" fill="{left_color}"/>
        <rect x="{right_x}" y="0" width="{right_w}" height="{d}" fill="{right_color}"/>
      </g>
      <circle cx="{r}" cy="{r}" r="{rr}" fill="none" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
    </svg>
    """.strip()

def solid_diamond_svg(size: int, color: str, border: str = NODE_BORDER_COLOR) -> str:
    # Manhole cover: circle with lugs
    center = size / 2
    outer_r = size * 0.40
    lug_size = size * 0.25
    gap = 2
    
    return f"""
    <svg xmlns="http://www.w3.org/2000/svg" width="{size}" height="{size}" viewBox="0 0 {size} {size}">
      <defs>
        <filter id="shadow" x="-50%" y="-50%" width="200%" height="200%">
          <feDropShadow dx="1" dy="1" stdDeviation="1" flood-opacity="0.4"/>
        </filter>
      </defs>
      
      <!-- Lugs (square handles) with shadow -->
      <rect x="{center - lug_size/2}" y="{gap}" width="{lug_size}" height="{lug_size}" fill="{color}" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}" filter="url(#shadow)"/>
      <rect x="{center - lug_size/2}" y="{size - lug_size - gap}" width="{lug_size}" height="{lug_size}" fill="{color}" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}" filter="url(#shadow)"/>
      <rect x="{gap}" y="{center - lug_size/2}" width="{lug_size}" height="{lug_size}" fill="{color}" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}" filter="url(#shadow)"/>
      <rect x="{size - lug_size - gap}" y="{center - lug_size/2}" width="{lug_size}" height="{lug_size}" fill="{color}" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}" filter="url(#shadow)"/>
      
      <!-- Central circle -->
      <circle cx="{center}" cy="{center}" r="{outer_r}" fill="{color}"/>
      <circle cx="{center}" cy="{center}" r="{outer_r}" fill="none" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
    </svg>
    """.strip()

def split_diamond_svg(size: int, left_color: str, right_color: str, border: str = NODE_BORDER_COLOR) -> str:
    # Manhole cover: circle with lugs (split colors)
    center = size / 2
    outer_r = size * 0.40
    lug_size = size * 0.25
    gap = 2
    
    return f"""
    <svg xmlns="http://www.w3.org/2000/svg" width="{size}" height="{size}" viewBox="0 0 {size} {size}">
      <defs>
        <filter id="shadowSplit" x="-50%" y="-50%" width="200%" height="200%">
          <feDropShadow dx="1" dy="1" stdDeviation="1" flood-opacity="0.4"/>
        </filter>
        <clipPath id="clipLugTop">
          <rect x="{center - lug_size/2}" y="{gap}" width="{lug_size}" height="{lug_size}"/>
        </clipPath>
        <clipPath id="clipLugBottom">
          <rect x="{center - lug_size/2}" y="{size - lug_size - gap}" width="{lug_size}" height="{lug_size}"/>
        </clipPath>
        <clipPath id="clipLugLeft">
          <rect x="{gap}" y="{center - lug_size/2}" width="{lug_size}" height="{lug_size}"/>
        </clipPath>
        <clipPath id="clipLugRight">
          <rect x="{size - lug_size - gap}" y="{center - lug_size/2}" width="{lug_size}" height="{lug_size}"/>
        </clipPath>
      </defs>
      
      <!-- Lugs with split colors -->
      <g clip-path="url(#clipLugTop)" filter="url(#shadowSplit)">
        <rect x="{center - lug_size/2}" y="{gap}" width="{lug_size/2}" height="{lug_size}" fill="{left_color}" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
        <rect x="{center}" y="{gap}" width="{lug_size/2}" height="{lug_size}" fill="{right_color}" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
      </g>
      <g clip-path="url(#clipLugBottom)" filter="url(#shadowSplit)">
        <rect x="{center - lug_size/2}" y="{size - lug_size - gap}" width="{lug_size/2}" height="{lug_size}" fill="{left_color}" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
        <rect x="{center}" y="{size - lug_size - gap}" width="{lug_size/2}" height="{lug_size}" fill="{right_color}" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
      </g>
      <g clip-path="url(#clipLugLeft)" filter="url(#shadowSplit)">
        <rect x="{gap}" y="{center - lug_size/2}" width="{lug_size}" height="{lug_size/2}" fill="{left_color}" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
        <rect x="{gap}" y="{center}" width="{lug_size}" height="{lug_size/2}" fill="{right_color}" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
      </g>
      <g clip-path="url(#clipLugRight)" filter="url(#shadowSplit)">
        <rect x="{size - lug_size - gap}" y="{center - lug_size/2}" width="{lug_size}" height="{lug_size/2}" fill="{left_color}" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
        <rect x="{size - lug_size - gap}" y="{center}" width="{lug_size}" height="{lug_size/2}" fill="{right_color}" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
      </g>
      
      <!-- Central circle with split colors -->
      <defs>
        <clipPath id="clipCircleLeft">
          <rect x="0" y="0" width="{center}" height="{size}"/>
        </clipPath>
        <clipPath id="clipCircleRight">
          <rect x="{center}" y="0" width="{center}" height="{size}"/>
        </clipPath>
      </defs>
      <circle cx="{center}" cy="{center}" r="{outer_r}" fill="{left_color}" clip-path="url(#clipCircleLeft)"/>
      <circle cx="{center}" cy="{center}" r="{outer_r}" fill="{right_color}" clip-path="url(#clipCircleRight)"/>
      <circle cx="{center}" cy="{center}" r="{outer_r}" fill="none" stroke="{border}" stroke-width="{SVG_BORDER_WIDTH}"/>
    </svg>
    """.strip()

def noise_color_kind(levels: set[int]) -> str:
    if levels == {1}:
        return "nl1"
    if levels == {2}:
        return "nl2"
    if levels == {3} or levels == {4}:
        return "nl34"
    if (1 in levels) and (2 in levels):
        return "mixed12"
    return "other"

def build_adjacency(connections_df: pd.DataFrame) -> dict[str, set[str]]:
    adj = defaultdict(set)
    if connections_df is None or connections_df.empty:
        return adj
    if "From" not in connections_df.columns or "To" not in connections_df.columns:
        return adj
    for _, r in connections_df.iterrows():
        a = r.get("From", None)
        b = r.get("To", None)
        if pd.isna(a) or pd.isna(b):
            continue
        a = str(a).strip()
        b = str(b).strip()
        if not a or not b:
            continue
        adj[a].add(b)
        adj[b].add(a)
    return adj

def neighborhood_nodes(adj: dict[str, set[str]], start: str, depth: int) -> set[str]:
    start = str(start).strip()
    if not start:
        return set()
    seen = {start}
    frontier = {start}
    for _ in range(max(0, int(depth))):
        nxt = set()
        for u in frontier:
            nxt |= set(adj.get(u, set()))
        nxt -= seen
        seen |= nxt
        frontier = nxt
        if not frontier:
            break
    return seen

def build_vis_nodes_edges(
    tray_df: pd.DataFrame,
    connections_df: pd.DataFrame,
    focus_node: str | None = None,
    focus_depth: int = 2,
    include_probe: bool = False,
    highlight_nodes: set[str] | None = None,
):
    tray_df = ensure_xy_columns(tray_df)
    tray_df = ensure_person_columns(tray_df)
    highlight_nodes = set(highlight_nodes or [])

    focus_set = None
    if focus_node:
        adj = build_adjacency(connections_df)
        focus_set = neighborhood_nodes(adj, focus_node, focus_depth)

    nodes = []
    for _, r in tray_df.iterrows():
        rn = r.get("RunName", None)
        if pd.isna(rn) or str(rn).strip() == "":
            continue
        rn = str(rn).strip()

        if focus_set is not None and rn not in focus_set:
            continue

        is_focus = (focus_node is not None and rn == str(focus_node).strip())
        is_highlight = rn in highlight_nodes

        levels = CableNetwork._parse_noise_levels(r.get("Noise Level", None), run_name=rn)
        kind = noise_color_kind(levels)
        email = str(r.get("Email", "") or "").strip()
        phone = str(r.get("Phone", "") or "").strip()
        role = str(r.get("Role", "") or "").strip()
        notes = str(r.get("Notes", "") or "").strip()

        if kind == "nl1":
            svg = person_circle_svg(84, ORANGE, NODE_BORDER_COLOR)
        elif kind == "nl2":
            svg = person_circle_svg(84, GREEN, NODE_BORDER_COLOR)
        elif kind == "nl34":
            svg = person_circle_svg(84, YELLOW, NODE_BORDER_COLOR)
        elif kind == "mixed12":
            svg = person_circle_svg(84, "#72BF44", NODE_BORDER_COLOR)
        else:
            svg = person_circle_svg(84, GRAY, NODE_BORDER_COLOR)

        node = {
            "id": rn,
            "label": rn,
            "title": (
                f"<b>{rn}</b><br/>Person node<br/>{noise_title(rn, r.get('Noise Level', ''))}"
                + (f"<br/>Role: {role}" if role else "")
                + (f"<br/>Email: {email}" if email else "")
                + (f"<br/>Phone: {phone}" if phone else "")
                + (f"<br/>Notes: {notes}" if notes else "")
            ),
            "shape": "image",
            "image": svg_data_uri(svg),
            "size": IMAGE_SIZE,
            "borderWidth": 0,
            "font": {"vadjust": 0},
        }

        if is_highlight:
            node["borderWidth"] = HIGHLIGHT_BORDER_WIDTH
            node["color"] = {"border": HIGHLIGHT_BORDER_COLOR}
            node["shadow"] = {"enabled": True, "size": HIGHLIGHT_GLOW_SIZE, "x": 0, "y": 0}
            node["size"] = IMAGE_SIZE + 18
            node["font"] = {"vadjust": 0, "size": 20}

        if is_focus:
            node["borderWidth"] = max(node.get("borderWidth", 0), 3)
            node["color"] = {"border": "#000000"}

        x = r.get("X", pd.NA)
        y = r.get("Y", pd.NA)
        try:
            x = None if pd.isna(x) else float(x)
        except Exception:
            x = None
        try:
            y = None if pd.isna(y) else float(y)  # Negate Y so increasing Y moves nodes up
        except Exception:
            y = None
        if x is not None and y is not None:
            node["x"] = x
            node["y"] = y

        nodes.append(node)

    node_ids = {n["id"] for n in nodes}

    edges = []
    seen = set()
    if connections_df is not None and not connections_df.empty and "From" in connections_df.columns and "To" in connections_df.columns:
        for _, r in connections_df.iterrows():
            a = r.get("From", None)
            b = r.get("To", None)
            if pd.isna(a) or pd.isna(b):
                continue
            a = str(a).strip()
            b = str(b).strip()
            if not a or not b:
                continue
            if a not in node_ids or b not in node_ids:
                continue
            key = tuple(sorted((a, b)))
            if key in seen:
                continue
            seen.add(key)
            
            edge_data = {
                "id": f"{a}|||{b}",
                "from": a,
                "to": b,
                "title": f"{a} ↔ {b}",
                "color": EDGE_COLOR,
                "width": EDGE_WIDTH,
            }
            
            # Add "EXP" label if exposed
            exposed_val = r.get("Exposed?", "")
            if str(exposed_val).strip().lower() == "yes":
                edge_data["label"] = "EXP"
                edge_data["font"] = {"size": 18, "align": "middle", "color": "red"}
                edge_data["labelHighlightBold"] = True
            
            edges.append(edge_data)

    if include_probe and PROBE_ID not in node_ids:
        nodes.append({
            "id": PROBE_ID,
            "label": "",
            "title": "",
            "shape": "dot",
            "size": 1,
            "x": 0,
            "y": 0,
            "fixed": True,
            "physics": False,
            "color": {"background": "rgba(0,0,0,0)", "border": "rgba(0,0,0,0)"},
            "font": {"size": 0, "color": "rgba(0,0,0,0)"},
            "selected": True,
        })

    return nodes, edges

def apply_positions_to_tray(tray_df: pd.DataFrame, positions: dict) -> pd.DataFrame:
    tray_df = ensure_xy_columns(tray_df)
    if not positions:
        return tray_df

    tray_df = tray_df.copy()
    rn_series = tray_df["RunName"].astype(str).str.strip()

    for nid, xy in positions.items():
        if str(nid).strip() == PROBE_ID:
            continue

        x = y = None
        if isinstance(xy, dict):
            x = xy.get("x")
            y = xy.get("y")
        elif isinstance(xy, (list, tuple)) and len(xy) >= 2:
            x, y = xy[0], xy[1]

        try:
            x = float(x)
            y = float(y)  # Negate Y when saving back
        except Exception:
            continue

        mask = rn_series == str(nid).strip()
        if mask.any():
            tray_df.loc[mask, "X"] = x
            tray_df.loc[mask, "Y"] = y

    return tray_df


def apply_offset_to_nodes(tray_df: pd.DataFrame, node_names: list[str], dx: float, dy: float) -> pd.DataFrame:
    """Offset specified nodes by dx, dy. Creates X/Y columns if missing."""
    tray_df = ensure_xy_columns(tray_df)
    tray_df = tray_df.copy()
    rn_series = tray_df["RunName"].astype(str).str.strip()
    
    for node_name in node_names:
        node_name = str(node_name).strip()
        if not node_name or node_name == PROBE_ID:
            continue
        
        mask = rn_series == node_name
        if mask.any():
            for idx in tray_df[mask].index:
                x_val = tray_df.loc[idx, "X"]
                y_val = tray_df.loc[idx, "Y"]
                
                # Convert to float, treating NaN as 0
                try:
                    x = float(x_val) if pd.notna(x_val) else 0.0
                except Exception:
                    x = 0.0
                try:
                    y = float(y_val) if pd.notna(y_val) else 0.0
                except Exception:
                    y = 0.0
                
                # Apply offset (negate dy so positive values move up)
                tray_df.loc[idx, "X"] = x + dx
                tray_df.loc[idx, "Y"] = y + dy
    
    return tray_df


def apply_scale_to_nodes(tray_df: pd.DataFrame, node_names: list[str], scale_factor: float) -> pd.DataFrame:
    """Scale specified nodes from their center point."""
    if not node_names or scale_factor <= 0:
        return tray_df
    
    tray_df = ensure_xy_columns(tray_df)
    tray_df = tray_df.copy()
    
    # Convert X and Y columns to float64 to avoid dtype warnings
    tray_df["X"] = pd.to_numeric(tray_df["X"], errors="coerce")
    tray_df["Y"] = pd.to_numeric(tray_df["Y"], errors="coerce")
    
    rn_series = tray_df["RunName"].astype(str).str.strip()
    
    # Calculate center point (average of all selected nodes)
    positions = []
    indices = []
    for node_name in node_names:
        node_name = str(node_name).strip()
        if not node_name or node_name == PROBE_ID:
            continue
        
        mask = rn_series == node_name
        if mask.any():
            for idx in tray_df[mask].index:
                x_val = tray_df.loc[idx, "X"]
                y_val = tray_df.loc[idx, "Y"]
                
                try:
                    x = float(x_val) if pd.notna(x_val) else 0.0
                except Exception:
                    x = 0.0
                try:
                    y = float(y_val) if pd.notna(y_val) else 0.0
                except Exception:
                    y = 0.0
                
                positions.append((x, y))
                indices.append(idx)
    
    if not positions:
        return tray_df
    
    # Calculate center
    center_x = sum(p[0] for p in positions) / len(positions)
    center_y = sum(p[1] for p in positions) / len(positions)
    
    # Scale each node from the center
    for i, idx in enumerate(indices):
        x, y = positions[i]
        
        # Calculate position relative to center
        rel_x = x - center_x
        rel_y = y - center_y
        
        # Scale the relative position
        new_rel_x = rel_x * scale_factor
        new_rel_y = rel_y * scale_factor
        
        # Calculate new absolute position
        tray_df.loc[idx, "X"] = center_x + new_rel_x
        tray_df.loc[idx, "Y"] = center_y + new_rel_y
    
    return tray_df


# ============================================================
# UNSAVED LAYOUT REMINDER helpers
# ============================================================

def _tray_xy_map(tray_df: pd.DataFrame) -> dict[str, tuple[float | None, float | None]]:
    df = ensure_xy_columns(tray_df)
    out: dict[str, tuple[float | None, float | None]] = {}
    if df is None or df.empty or "RunName" not in df.columns:
        return out

    for _, r in df.iterrows():
        rn = r.get("RunName", None)
        if pd.isna(rn) or str(rn).strip() == "":
            continue
        rn = str(rn).strip()

        x = r.get("X", pd.NA)
        y = r.get("Y", pd.NA)
        try:
            x = None if pd.isna(x) else float(x)
        except Exception:
            x = None
        try:
            y = None if pd.isna(y) else float(y)
        except Exception:
            y = None
        out[rn] = (x, y)
    return out

def _positions_map(positions: dict) -> dict[str, tuple[float | None, float | None]]:
    out: dict[str, tuple[float | None, float | None]] = {}
    if not isinstance(positions, dict):
        return out
    for nid, xy in positions.items():
        if str(nid).strip() == PROBE_ID:
            continue

        x = y = None
        if isinstance(xy, dict):
            x = xy.get("x")
            y = xy.get("y")
        elif isinstance(xy, (list, tuple)) and len(xy) >= 2:
            x, y = xy[0], xy[1]

        try:
            x = float(x)
            y = float(y)
        except Exception:
            x = y = None

        out[str(nid).strip()] = (x, y)
    return out

def _has_unsaved_layout_changes(tray_df: pd.DataFrame, last_positions: dict | None, tol: float = 0.5) -> bool:
    if not isinstance(last_positions, dict) or not last_positions:
        return False

    tmap = _tray_xy_map(tray_df)
    pmap = _positions_map(last_positions)

    if not pmap:
        return False

    for nid, (px, py) in pmap.items():
        tx, ty = tmap.get(nid, (None, None))
        if px is None or py is None or tx is None or ty is None:
            continue
        if abs(px - tx) > tol or abs(py - ty) > tol:
            return True

    return False


# ============================================================
# DataFrame mutation helpers
# ============================================================

def df_delete_node(tray_df, connections_df, endpoints_df, node: str):
    node = str(node).strip()
    tray_df2 = tray_df.copy()
    tray_df2 = tray_df2[tray_df2["RunName"].astype(str).str.strip() != node].reset_index(drop=True)

    con2 = connections_df.copy()
    if "From" in con2.columns and "To" in con2.columns:
        con2 = con2[
            (con2["From"].astype(str).str.strip() != node) &
            (con2["To"].astype(str).str.strip() != node)
        ].reset_index(drop=True)

    ep2 = endpoints_df.copy()
    if "tray/conduit(s)" in ep2.columns:
        def remove_token(s):
            parts = [p.strip() for p in str(s).split(",") if p.strip()]
            parts = [p for p in parts if p != node]
            return ",".join(parts)
        ep2["tray/conduit(s)"] = ep2["tray/conduit(s)"].apply(remove_token)

    return tray_df2, con2, ep2

def df_delete_edge(connections_df, a: str, b: str):
    aa = str(a).strip()
    bb = str(b).strip()
    con = connections_df.copy()

    def is_match(r):
        x = str(r.get("From", "")).strip()
        y = str(r.get("To", "")).strip()
        return set([x, y]) == set([aa, bb])

    if ("From" in con.columns) and ("To" in con.columns):
        mask = con.apply(is_match, axis=1)
        con = con[~mask].reset_index(drop=True)
    return con

def df_add_edge(connections_df: pd.DataFrame, a: str, b: str) -> tuple[pd.DataFrame, bool]:
    a = str(a).strip()
    b = str(b).strip()
    if not a or not b or a == b:
        return connections_df, False

    con = connections_df.copy()

    if "From" not in con.columns or "To" not in con.columns:
        return connections_df, False

    def is_dup(r):
        x = str(r.get("From", "")).strip()
        y = str(r.get("To", "")).strip()
        return set([x, y]) == set([a, b])

    if not con.empty:
        try:
            if bool(con.apply(is_dup, axis=1).any()):
                return connections_df, False
        except Exception:
            pass

    new_row = {col: "" for col in con.columns}
    new_row["From"] = a
    new_row["To"] = b
    con = pd.concat([con, pd.DataFrame([new_row])], ignore_index=True)
    return con, True

def get_edge_exposed_status(connections_df: pd.DataFrame, a: str, b: str) -> bool:
    """Get whether an edge is marked as exposed."""
    a = str(a).strip()
    b = str(b).strip()
    
    if "From" not in connections_df.columns or "To" not in connections_df.columns:
        return False
    
    def is_match(r):
        x = str(r.get("From", "")).strip()
        y = str(r.get("To", "")).strip()
        return set([x, y]) == set([a, b])
    
    for _, row in connections_df.iterrows():
        if is_match(row):
            exposed_val = row.get("Exposed?", "")
            return str(exposed_val).strip().lower() == "yes"
    
    return False

def set_edge_exposed_status(connections_df: pd.DataFrame, a: str, b: str, exposed: bool) -> pd.DataFrame:
    """Set whether an edge is marked as exposed."""
    a = str(a).strip()
    b = str(b).strip()
    
    if "From" not in connections_df.columns or "To" not in connections_df.columns:
        return connections_df
    
    con = connections_df.copy()
    
    def is_match(r):
        x = str(r.get("From", "")).strip()
        y = str(r.get("To", "")).strip()
        return set([x, y]) == set([a, b])
    
    for idx, row in con.iterrows():
        if is_match(row):
            con.at[idx, "Exposed?"] = "yes" if exposed else ""
    
    return con

def _compute_default_xy_near_network(tray_df: pd.DataFrame) -> tuple[float | None, float | None]:
    df = ensure_xy_columns(tray_df)
    xs = pd.to_numeric(df["X"], errors="coerce")
    ys = pd.to_numeric(df["Y"], errors="coerce")
    mask = xs.notna() & ys.notna()
    if mask.any():
        cx = float(xs[mask].mean())
        cy = float(ys[mask].mean())
        return cx + 140.0, cy
    return 200.0, 200.0

def df_add_node(tray_df: pd.DataFrame, name: str, noise_level_text: str = "", x=None, y=None) -> tuple[pd.DataFrame, bool]:
    name = str(name).strip()
    if not name:
        return tray_df, False

    df = ensure_person_columns(ensure_xy_columns(tray_df.copy()))
    
    # Ensure X and Y columns are float64 dtype to avoid dtype warnings
    df["X"] = pd.to_numeric(df["X"], errors="coerce")
    df["Y"] = pd.to_numeric(df["Y"], errors="coerce")

    if "RunName" not in df.columns:
        df["RunName"] = ""
    if "Noise Level" not in df.columns:
        df["Noise Level"] = ""

    if (df["RunName"].astype(str).str.strip() == name).any():
        return tray_df, False

    new_row = {col: "" for col in df.columns}
    new_row["RunName"] = name

    nl_text = str(noise_level_text or "").strip()
    if nl_text:
        new_row["Noise Level"] = nl_text

    if (x is None or str(x).strip() == "") and (y is None or str(y).strip() == ""):
        dx, dy = _compute_default_xy_near_network(df)
        new_row["X"] = dx if dx is not None else None
        new_row["Y"] = dy if dy is not None else None
    else:
        try:
            if x is not None and str(x).strip() != "":
                new_row["X"] = float(x)
        except Exception:
            new_row["X"] = None
        try:
            if y is not None and str(y).strip() != "":
                new_row["Y"] = float(y)
        except Exception:
            new_row["Y"] = None

    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    return df, True

def df_rename_node(tray_df, connections_df, endpoints_df, old: str, new: str):
    old = str(old).strip()
    new = str(new).strip()
    if not old or not new or old == new:
        return tray_df, connections_df, endpoints_df

    t = tray_df.copy()
    mask = t["RunName"].astype(str).str.strip() == old
    if mask.any():
        t.loc[mask, "RunName"] = new

    c = connections_df.copy()
    if "From" in c.columns:
        c.loc[c["From"].astype(str).str.strip() == old, "From"] = new
    if "To" in c.columns:
        c.loc[c["To"].astype(str).str.strip() == old, "To"] = new

    e = endpoints_df.copy()
    if "tray/conduit(s)" in e.columns:
        def rename_token(s):
            parts = [p.strip() for p in str(s).split(",") if p.strip()]
            parts = [new if p == old else p for p in parts]
            return ",".join(parts)
        e["tray/conduit(s)"] = e["tray/conduit(s)"].apply(rename_token)

    return t, c, e

def df_duplicate_node(tray_df, source_name: str, new_name: str):
    source_name = str(source_name).strip()
    new_name = str(new_name).strip()
    if not source_name or not new_name:
        return tray_df
    df = ensure_xy_columns(tray_df.copy())

    if (df["RunName"].astype(str).str.strip() == new_name).any():
        return df

    src_mask = df["RunName"].astype(str).str.strip() == source_name
    if not src_mask.any():
        return df

    src_row = df[src_mask].iloc[0].to_dict()
    src_row["RunName"] = new_name

    try:
        if not pd.isna(src_row.get("X")):
            src_row["X"] = float(src_row["X"]) + 50.0
        if not pd.isna(src_row.get("Y")):
            src_row["Y"] = float(src_row["Y"]) + 50.0
    except Exception:
        pass

    df = pd.concat([df, pd.DataFrame([src_row])], ignore_index=True)
    return df

def df_duplicate_nodes(tray_df: pd.DataFrame, source_names: list[str]) -> tuple[pd.DataFrame, list[str]]:
    """Duplicate multiple nodes by adding _COPY suffix to each name. Returns updated df and list of new names."""
    df = ensure_xy_columns(tray_df.copy())
    new_names = []
    
    for source_name in source_names:
        source_name = str(source_name).strip()
        if not source_name:
            continue
        
        # Generate new name with _COPY suffix
        new_name = f"{source_name}_COPY"
        counter = 1
        while (df["RunName"].astype(str).str.strip() == new_name).any():
            new_name = f"{source_name}_COPY_{counter}"
            counter += 1
        
        src_mask = df["RunName"].astype(str).str.strip() == source_name
        if not src_mask.any():
            continue
        
        src_row = df[src_mask].iloc[0].to_dict()
        src_row["RunName"] = new_name
        
        try:
            if not pd.isna(src_row.get("X")):
                src_row["X"] = float(src_row["X"]) + 50.0
            if not pd.isna(src_row.get("Y")):
                src_row["Y"] = float(src_row["Y"]) + 50.0
        except Exception:
            pass
        
        df = pd.concat([df, pd.DataFrame([src_row])], ignore_index=True)
        new_names.append(new_name)
    
    return df, new_names

def df_set_node_noise_level(tray_df: pd.DataFrame, node: str, noise_level_text: str) -> tuple[pd.DataFrame, bool]:
    node = str(node).strip()
    df = tray_df.copy()
    if "RunName" not in df.columns:
        return tray_df, False
    if "Noise Level" not in df.columns:
        df["Noise Level"] = ""

    mask = df["RunName"].astype(str).str.strip() == node
    if not mask.any():
        return tray_df, False

    df.loc[mask, "Noise Level"] = str(noise_level_text or "").strip()
    return df, True


# ============================================================
# Selection parsing helper (edge delete)
# ============================================================

def parse_selected_edge(sel_edge) -> tuple[str | None, str | None, str]:
    if sel_edge is None:
        return None, None, "None"

    if isinstance(sel_edge, dict):
        a = str(sel_edge.get("from", "")).strip()
        b = str(sel_edge.get("to", "")).strip()
        if a and b:
            return a, b, f"{a} ↔ {b}"
        sid = sel_edge.get("id")
        if isinstance(sid, str) and "|||" in sid:
            x, y = sid.split("|||", 1)
            return x.strip(), y.strip(), f"{x.strip()} ↔ {y.strip()}"
        return None, None, str(sel_edge)

    if isinstance(sel_edge, (list, tuple)):
        if len(sel_edge) >= 2:
            a = str(sel_edge[0]).strip()
            b = str(sel_edge[1]).strip()
            if a and b:
                return a, b, f"{a} ↔ {b}"
        return None, None, str(sel_edge)

    if isinstance(sel_edge, str):
        s = sel_edge.strip()
        if "|||" in s:
            x, y = s.split("|||", 1)
            return x.strip(), y.strip(), f"{x.strip()} ↔ {y.strip()}"
        return None, None, s

    return None, None, str(sel_edge)


# ============================================================
# Generate route helpers (combined search + warnings)
# ============================================================

def _autoroute_pack(kind: str, value: str) -> str:
    k = str(kind or "").strip().upper()
    v = str(value or "").strip()
    return f"{k}::{v}"

def _autoroute_unpack(packed: str) -> tuple[str, str]:
    s = str(packed or "").strip()
    if "::" not in s:
        return "", s
    k, v = s.split("::", 1)
    return k.strip().upper(), v.strip()

def _format_autoroute_option(packed: str) -> str:
    k, v = _autoroute_unpack(packed)
    if not v:
        return ""
    if k == "EP":
        return f"Endpoint: {v}"
    if k == "TRAY":
        t = infer_type_from_name(v)
        if t == "Tray":
            return f"Tray: {v}"
        if t == "Conduit":
            return f"Conduit: {v}"
        return f"Node: {v}"
    return v

def _resolve_autoroute_side_to_trays(
    net: CableNetwork,
    kind: str,
    value: str,
    noise_level: int,
) -> tuple[list[str], str]:
    k = str(kind or "").strip().upper()
    val = str(value or "").strip().lstrip("+")
    nl = int(noise_level)

    if not val:
        return [], "Nothing selected."

    if k == "TRAY":
        if val not in net.noise_levels:
            return [], f"{infer_type_from_name(val)} '{val}' is not present in Tray.RunName."
        levels = net.noise_levels.get(val, set())
        if not levels:
            return [], f"{infer_type_from_name(val)} '{val}' has no parsed noise level (Tray.Noise Level is blank/unparseable)."
        if nl not in levels:
            return [], f"{infer_type_from_name(val)} '{val}' does not support Noise Level {nl} (it has {sorted(levels)})."
        return [val], ""

    if val not in net.endpoints:
        return [], f"Endpoint '{val}' is not present in Endpoints.device/panel."
    all_trays = [str(t).strip() for t in (net.endpoints.get(val, []) or []) if str(t).strip()]
    if not all_trays:
        return [], f"Endpoint '{val}' has no tray/conduit(s) listed in Endpoints.tray/conduit(s)."

    missing = [t for t in all_trays if t not in net.noise_levels]
    present = [t for t in all_trays if t in net.noise_levels]

    if not present and missing:
        return [], (
            f"Endpoint '{val}' references trays/conduits that are not present in Tray.RunName: "
            + ", ".join(sorted(set(missing)))
        )

    matches = [t for t in present if nl in (net.noise_levels.get(t, set()) or set())]
    if not matches:
        parts = []
        if present:
            parts.append(
                "No referenced trays/conduits support Noise Level "
                f"{nl} (present: {', '.join(sorted(set(present)))})"
            )
        if missing:
            parts.append(
                "Some referenced trays/conduits are missing from Tray.RunName "
                f"({', '.join(sorted(set(missing)))})"
            )
        return [], f"Endpoint '{val}': " + " | ".join(parts)

    return matches, ""

def try_auto_route_packed(
    tray_df: pd.DataFrame,
    connections_df: pd.DataFrame,
    endpoints_df: pd.DataFrame,
    start_packed: str,
    end_packed: str,
    noise_level: int,
) -> tuple[list[str] | None, str]:
    nl = int(noise_level) if noise_level is not None else None
    if nl is None:
        return None, "Pick a noise level."

    sk, sv = _autoroute_unpack(start_packed)
    ek, ev = _autoroute_unpack(end_packed)

    if not sv and not ev:
        return None, "Pick both start and end."
    if not sv:
        return None, "Pick a start."
    if not ev:
        return None, "Pick an end."

    net = CableNetwork()
    net.build_from_dfs(tray_df, connections_df, endpoints_df)

    start_trays, start_err = _resolve_autoroute_side_to_trays(net, sk, sv, nl)
    end_trays, end_err = _resolve_autoroute_side_to_trays(net, ek, ev, nl)

    if start_err and end_err:
        return None, f"Start issue: {start_err}\nEnd issue: {end_err}"
    if start_err:
        return None, f"Start issue: {start_err}"
    if end_err:
        return None, f"End issue: {end_err}"

    path = net.find_route(start_trays, set(end_trays), nl)
    if not path:
        start_desc = f"{len(start_trays)} start candidate(s): {', '.join(sorted(set(start_trays)))}"
        end_desc = f"{len(end_trays)} end candidate(s): {', '.join(sorted(set(end_trays)))}"
        return None, (
            "No valid route found in the graph at Noise Level "
            f"{nl}.\n"
            f"{start_desc}\n"
            f"{end_desc}\n"
            "Likely causes:\n"
            "- Missing/incorrect rows in Connections (graph disconnected)\n"
            "- Trays/conduits exist but routing is blocked by noise level constraints\n"
        )

    return path, "OK"

def validate_manual_route_steps(connections_df: pd.DataFrame, nodes: list[str]) -> tuple[bool, list[tuple[str, str]]]:
    if not nodes or len(nodes) < 2:
        return True, []
    adj = build_adjacency(connections_df)
    bad = []
    for a, b in zip(nodes[:-1], nodes[1:]):
        if b not in adj.get(a, set()):
            bad.append((a, b))
    return (len(bad) == 0), bad


# ============================================================
# Streamlit UI
# ============================================================

st.set_page_config(page_title="Cable Routing Webapp", layout="wide")

st.markdown(
    "<style>div.block-container {padding-top: 1rem;}</style>",
    unsafe_allow_html=True,
)

st.title("Cable Routing Webapp")

st.sidebar.header("Workbook")

st.session_state.setdefault("uploader_key_v", 0)
uploader_key = f"uploader_{st.session_state.uploader_key_v}"

source = st.sidebar.radio(
    "Choose workbook source",
    options=["Upload Excel (.xlsx)", "Use demo workbook"],
    index=0,
    key="workbook_source",
)

uploaded = None
demo_clicked = False

if source == "Upload Excel (.xlsx)":
    uploaded = st.sidebar.file_uploader("Upload Excel (.xlsx)", type=["xlsx"], key=uploader_key)
else:
    demo_clicked = st.sidebar.button("Load demo workbook", key="load_demo_btn", width="stretch")
    st.sidebar.caption("Loads a built-in example with multiple routes (Noise Levels 1 and 2 only).")
    demo_bytes_for_dl = build_demo_workbook_bytes()
    st.sidebar.download_button(
        "Download demo workbook (.xlsx)",
        data=demo_bytes_for_dl,
        file_name="demo_network_configuration.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.session_state.setdefault("tray_df", None)
st.session_state.setdefault("connections_df", None)
st.session_state.setdefault("endpoints_df", None)
st.session_state.setdefault("cables_df", None)
st.session_state.setdefault("routes_df", None)
st.session_state.setdefault("upload_hash", None)
st.session_state.setdefault("is_older_template", False)

st.session_state.setdefault("sel_nodes", [])
st.session_state.setdefault("sel_edges", [])

st.session_state.setdefault("focus_node", None)
st.session_state.setdefault("focus_depth", 2)

st.session_state.setdefault("graph_key_v", 0)

st.session_state.setdefault("layout_opt_active", False)
st.session_state.setdefault("layout_opt_last_positions", None)
st.session_state.setdefault("layout_opt_backup_xy", None)

st.session_state.setdefault("endpoint_highlight_nodes", set())
st.session_state.setdefault("endpoint_highlight_ep", None)

st.session_state.setdefault("route_highlight_nodes", set())
st.session_state.setdefault("route_highlight_cable", None)

st.session_state.setdefault("initial_layout_autosave_active", False)

st.session_state.setdefault("gen_route_mode", "Auto-route (endpoints)")
st.session_state.setdefault("gen_route_noise_level", 1)
st.session_state.setdefault("gen_route_auto_nodes", set())
st.session_state.setdefault("gen_route_auto_via", "")
st.session_state.setdefault("gen_route_auto_msg", "")

st.session_state.setdefault("gen_route_manual_nodes", [])
st.session_state.setdefault("gen_route_manual_via", "")
st.session_state.setdefault("gen_route_manual_last_clicked", None)

st.session_state.setdefault("layout_unsaved_hint", False)

# Save positions notice (must render BELOW the button)
st.session_state.setdefault("save_positions_notice", None)  # ("success"/"warning"/"info"/"error", "message")

# Track graph interactions so we can CLEAR the notice when user clicks/drags/changes graph
st.session_state.setdefault("graph_interaction_sig", None)

# Multi-node offset feature
st.session_state.setdefault("multi_node_offset_enabled", False)
st.session_state.setdefault("multi_node_offset_dx", 0.0)
st.session_state.setdefault("multi_node_offset_dy", 0.0)

GRAPH_HEIGHT = 750

st.markdown(
    """
    <style>
      div[data-testid="stVerticalBlockBorderWrapper"]{
        border-radius: 10px !important;
      }
      .graphwrap, .graphwrap > div { width: 100% !important; }
      .graphwrap iframe { width: 100% !important; min-width: 100% !important; display: block !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

if st.sidebar.button("Clear workbook", key="clear_workbook_btn"):
    st.session_state.tray_df = None
    st.session_state.connections_df = None
    st.session_state.endpoints_df = None
    st.session_state.cables_df = None
    st.session_state.routes_df = None
    st.session_state.upload_hash = None
    st.session_state.sel_nodes = []
    st.session_state.sel_edges = []
    st.session_state.focus_node = None
    st.session_state.graph_key_v += 1
    st.session_state.uploader_key_v += 1
    st.session_state.layout_opt_active = False
    st.session_state.layout_opt_last_positions = None
    st.session_state.layout_opt_backup_xy = None
    st.session_state.endpoint_highlight_nodes = set()
    st.session_state.endpoint_highlight_ep = None
    st.session_state.route_highlight_nodes = set()
    st.session_state.route_highlight_cable = None

    st.session_state.initial_layout_autosave_active = False

    st.session_state.gen_route_mode = "Auto-route (endpoints)"
    st.session_state.gen_route_noise_level = 1
    st.session_state.gen_route_auto_nodes = set()
    st.session_state.gen_route_auto_via = ""
    st.session_state.gen_route_auto_msg = ""
    st.session_state.gen_route_manual_nodes = []
    st.session_state.gen_route_manual_via = ""
    st.session_state.gen_route_manual_last_clicked = None

    st.session_state.layout_unsaved_hint = False
    st.session_state.save_positions_notice = None
    st.session_state.graph_interaction_sig = None

    st.rerun()

file_bytes = None
if uploaded is not None:
    file_bytes = uploaded.getvalue()
if demo_clicked:
    file_bytes = build_demo_workbook_bytes()

if file_bytes is not None:
    try:
        this_hash = hashlib.md5(file_bytes).hexdigest()
        if st.session_state.upload_hash != this_hash:
            loaded, is_older_template = load_excel_to_dfs(file_bytes)

            tdf_loaded = ensure_xy_columns(loaded["Tray"])
            st.session_state.tray_df = tdf_loaded
            st.session_state.connections_df = loaded["Connections"]
            st.session_state.endpoints_df = loaded["Endpoints"]
            st.session_state.cables_df = loaded["Cables(input)"]
            st.session_state.routes_df = None
            st.session_state.upload_hash = this_hash
            st.session_state.is_older_template = is_older_template
            st.session_state.sel_nodes = []
            st.session_state.sel_edges = []
            st.session_state.focus_node = None
            st.session_state.graph_key_v += 1
            st.session_state.layout_opt_active = False
            st.session_state.layout_opt_last_positions = None
            st.session_state.layout_opt_backup_xy = None
            st.session_state.endpoint_highlight_nodes = set()
            st.session_state.endpoint_highlight_ep = None
            st.session_state.route_highlight_nodes = set()
            st.session_state.route_highlight_cable = None

            st.session_state.initial_layout_autosave_active = (not tray_has_any_xy(tdf_loaded))

            st.session_state.gen_route_mode = "Auto-route (endpoints)"
            st.session_state.gen_route_noise_level = 1
            st.session_state.gen_route_auto_nodes = set()
            st.session_state.gen_route_auto_via = ""
            st.session_state.gen_route_auto_msg = ""
            st.session_state.gen_route_manual_nodes = []
            st.session_state.gen_route_manual_via = ""
            st.session_state.gen_route_manual_last_clicked = None

            st.session_state.layout_unsaved_hint = False
            st.session_state.save_positions_notice = None
            st.session_state.graph_interaction_sig = None

            if is_older_template:
                st.sidebar.warning(
                    "⚠️ **Older template detected** — Your workbook is using an older template without the INCLUDE/EXCLUDE columns. "
                    "These columns have been automatically added with empty values. "
                    "Once you route your cables and download the updated workbook, you'll have the latest template format with "
                    "support for custom route preferences."
                )
            st.sidebar.success("Workbook loaded.")
    except Exception as e:
        st.sidebar.error(f"Failed to load workbook: {e}")
        st.stop()

if st.session_state.tray_df is None:
    st.info("Upload an Excel workbook or load the demo workbook to begin.")
    st.stop()

st.session_state.tray_df = ensure_xy_columns(st.session_state.tray_df)
st.session_state.tray_df = ensure_person_columns(st.session_state.tray_df)

dfs = {
    "Tray": st.session_state.tray_df,
    "Connections": st.session_state.connections_df,
    "Endpoints": st.session_state.endpoints_df,
    "Cables(input)": st.session_state.cables_df,
}

val_errors = validate_dfs(dfs)
if val_errors:
    st.warning("Workbook validation warnings:")
    for e in val_errors:
        st.write(f"- {e}")

tab1, tab2, tab3, tab4, tabG, tab5, tab_reverse = st.tabs(
    ["📦 Tray", "🔗 Connections", "🎯 Endpoints", "🧵 Cables(input)", "🗺️ Graph Editor", "🧾 Routing Output", "🔁 Reverse Engineer"]
)

with tab1:
    st.subheader("Tray")
    st.caption("X and Y store node positions. Email, phone, role, and notes store person-style metadata for each node.")
    st.session_state.tray_df = st.data_editor(
        ensure_person_columns(st.session_state.tray_df),
        num_rows="dynamic",
        width="stretch",
        key="tray_editor",
        column_config={
            "Email": st.column_config.TextColumn("Email"),
            "Phone": st.column_config.TextColumn("Phone"),
            "Role": st.column_config.TextColumn("Role"),
            "Notes": st.column_config.TextColumn("Notes", width="large"),
        },
    )

with tab2:
    st.subheader("Connections")
    st.caption("The 'Exposed?' column allows marking edges as exposed (yes/no or blank).")
    st.session_state.connections_df = st.data_editor(
        st.session_state.connections_df,
        num_rows="dynamic",
        width="stretch",
        key="connections_editor",
    )

with tab3:
    st.subheader("Endpoints")
    st.caption("Column C (3rd column) is interpreted as exposed yes/no (case-insensitive).")
    st.session_state.endpoints_df = st.data_editor(
        st.session_state.endpoints_df,
        num_rows="dynamic",
        width="stretch",
        key="endpoints_editor",
    )

with tab4:
    st.subheader("Cables(input)")
    st.session_state.cables_df = st.data_editor(
        st.session_state.cables_df,
        num_rows="dynamic",
        width="stretch",
        key="cables_editor",
    )

with tabG:

    graph_col, tools_col = st.columns([3, 1], gap="large")

    node_names = (
        st.session_state.tray_df["RunName"]
        .dropna()
        .astype(str)
        .map(str.strip)
        .replace("", pd.NA)
        .dropna()
        .unique()
        .tolist()
    )
    node_names = sorted(node_names, key=lambda s: s.lower())

    endpoint_to_trays = build_endpoint_to_trays_map(st.session_state.endpoints_df)
    tray_to_endpoints = build_tray_to_endpoints_map(st.session_state.endpoints_df)
    endpoint_names = sorted(endpoint_to_trays.keys(), key=lambda s: s.lower())

    combined_options = [""] + (
        [_autoroute_pack("EP", ep) for ep in endpoint_names] +
        [_autoroute_pack("TRAY", n) for n in node_names]
    )

    selection = None
    this_run_sig = None

    with graph_col:
        graph_box = st.container(border=False)
        with graph_box:
            st.markdown('<div class="graphwrap">', unsafe_allow_html=True)

            combined_highlight = (
                set(st.session_state.endpoint_highlight_nodes or set())
                | set(st.session_state.route_highlight_nodes or set())
                | set(st.session_state.gen_route_auto_nodes or set())
                | set(st.session_state.gen_route_manual_nodes or [])
            )

            include_probe = bool(st.session_state.layout_opt_active or st.session_state.initial_layout_autosave_active)

            nodes, edges = build_vis_nodes_edges(
                st.session_state.tray_df,
                st.session_state.connections_df,
                focus_node=st.session_state.focus_node,
                focus_depth=int(st.session_state.focus_depth),
                include_probe=include_probe,
                highlight_nodes=combined_highlight,
            )

            if st.session_state.layout_opt_active:
                options = {
                    "physics": {
                        "enabled": True,
                        "stabilization": {
                            "enabled": True,
                            "iterations": 1600,
                            "updateInterval": 50,
                            "fit": False,
                        },
                        "solver": "forceAtlas2Based",
                        "forceAtlas2Based": {
                            "gravitationalConstant": -260,
                            "centralGravity": 0.01,
                            "springLength": 240,
                            "springConstant": 0.02,
                            "damping": 0.4,
                            "avoidOverlap": 1.0,
                        },
                        "minVelocity": 0.20,
                        "maxVelocity": 70,
                        "timestep": 0.5,
                    },
                    "layout": {"improvedLayout": True},
                    "interaction": {
                        "dragNodes": True,
                        "dragView": True,
                        "zoomView": True,
                        "hover": True,
                        "multiselect": True,
                        "selectConnectedEdges": True,
                        "navigationButtons": False,
                        "keyboard": False,
                    },
                    "nodes": {"font": {"vadjust": 0}},
                    "edges": {"smooth": False, "color": EDGE_COLOR, "width": EDGE_WIDTH},
                    # "animation": {"duration": 0},
                }
            else:
                options = {
                    "physics": {"enabled": False},
                    "layout": {"improvedLayout": True},
                    "interaction": {
                        "dragNodes": True,
                        "dragView": True,
                        "zoomView": True,
                        "hover": True,
                        "multiselect": True,
                        "selectConnectedEdges": True,
                        "navigationButtons": False,
                        "keyboard": False,
                    },
                    "nodes": {"font": {"vadjust": 0}},
                    "edges": {"smooth": False, "color": EDGE_COLOR, "width": EDGE_WIDTH},
                    # "animation": {"duration": 0},
                }

            selection = streamlit_vis_network(
                nodes,
                edges,
                height=GRAPH_HEIGHT,
                options=options,
                key=f"vis_network_graph_{st.session_state.graph_key_v}",
            )

            st.markdown("</div>", unsafe_allow_html=True)

        clicked_node = None
        positions = None
        sel_nodes = []
        sel_edges = []

        if selection:
            try:
                sel_nodes, sel_edges, positions = selection
            except Exception:
                sel_nodes, sel_edges, positions = [], [], None

            st.session_state.layout_opt_last_positions = positions if isinstance(positions, dict) else st.session_state.layout_opt_last_positions

            cleaned_nodes = [n for n in (sel_nodes or []) if str(n).strip() != PROBE_ID]
            st.session_state.sel_nodes = cleaned_nodes
            st.session_state.sel_edges = sel_edges or []

            if cleaned_nodes:
                clicked_node = str(cleaned_nodes[0]).strip()
        else:
            sel_nodes, sel_edges, positions = [], [], None
            st.session_state.sel_nodes = []
            st.session_state.sel_edges = []

        def _positions_sig(pos: dict | None):
            if not isinstance(pos, dict) or not pos:
                return (0, 0, 0)
            items = [(k, v) for k, v in pos.items() if str(k).strip() != PROBE_ID]
            n = len(items)
            sx = sy = 0.0
            for k, v in items[:15]:
                x = y = 0.0
                if isinstance(v, dict):
                    x = float(v.get("x", 0.0) or 0.0)
                    y = float(v.get("y", 0.0) or 0.0)
                elif isinstance(v, (list, tuple)) and len(v) >= 2:
                    try:
                        x = float(v[0] or 0.0)
                        y = float(v[1] or 0.0)
                    except Exception:
                        x = y = 0.0
                sx += x
                sy += y
            return (n, round(sx, 1), round(sy, 1))

        this_run_sig = (
            tuple([str(x).strip() for x in (sel_nodes or []) if str(x).strip() != PROBE_ID]),
            tuple([str(x) for x in (sel_edges or [])][:3]),
            _positions_sig(positions),
        )

        # Clear message on real graph interaction, BUT not immediately after a "Save" forced refresh.
        if st.session_state.graph_interaction_sig is None:
            st.session_state.graph_interaction_sig = this_run_sig
        else:
            if this_run_sig != st.session_state.graph_interaction_sig:
                st.session_state.graph_interaction_sig = this_run_sig
                st.session_state.save_positions_notice = None

        if (not st.session_state.layout_opt_active) and isinstance(st.session_state.layout_opt_last_positions, dict):
            st.session_state.layout_unsaved_hint = _has_unsaved_layout_changes(
                st.session_state.tray_df,
                st.session_state.layout_opt_last_positions,
                tol=0.5,
            )
        else:
            st.session_state.layout_unsaved_hint = False

        if (
            st.session_state.initial_layout_autosave_active
            and (not st.session_state.layout_opt_active)
            and isinstance(positions, dict)
            and len(positions) > 0
        ):
            st.session_state.tray_df = apply_positions_to_tray(st.session_state.tray_df, positions)
            st.session_state.initial_layout_autosave_active = False
            st.session_state.graph_key_v += 1
            st.success("Initial layout locked in and saved to Tray (X/Y).")
            st.rerun()

        if (
            st.session_state.gen_route_mode == "Manual (click nodes in graph)"
            and clicked_node
            and clicked_node in node_names
        ):
            last = st.session_state.gen_route_manual_last_clicked
            if clicked_node != last:
                seq = list(st.session_state.gen_route_manual_nodes or [])
                if not seq or seq[-1] != clicked_node:
                    seq.append(clicked_node)
                    st.session_state.gen_route_manual_nodes = seq
                    st.session_state.gen_route_manual_via = format_route_nodes_for_via(
                        seq,
                        noise_level=int(st.session_state.gen_route_noise_level) if st.session_state.gen_route_noise_level is not None else None,
                    )
                st.session_state.gen_route_manual_last_clicked = clicked_node
                st.session_state.graph_key_v += 1
                st.rerun()

    with tools_col:
        # Save button (OUTSIDE scrollable container - stays visible when scrolling)
        save_clicked = st.button("💾 Save current node positions to Tray (X/Y)", key="save_positions_btn", width="stretch")
        save_notice_slot = st.empty()
        
        # Scrollable tools container
        tools_box = st.container(height=GRAPH_HEIGHT, border=False)
        with tools_box:
            if st.session_state.layout_unsaved_hint and (not st.session_state.layout_opt_active):
                st.warning(
                    "You have **unsaved node position changes**.\n\n"
                    "Click **Save current node positions to Tray (X/Y)** to store the new layout in the workbook."
                )

            st.markdown("### Selection")

            sel_nodes2 = st.session_state.sel_nodes or []
            sel_edges2 = st.session_state.sel_edges or []

            # Multi-node operations
            if len(sel_nodes2) > 1:
                # Connect all selected nodes (FIRST)
                st.markdown("**Connect All Selected Nodes**")
                st.caption(f"Create connections between all {len(sel_nodes2)} selected nodes")
                
                if st.button("Connect all selected nodes", key="connect_all_nodes_btn", width="stretch"):
                    connections_added = 0
                    connections_already_exist = 0
                    
                    # Create connections between all pairs of selected nodes
                    for i, node1 in enumerate(sel_nodes2):
                        for node2 in sel_nodes2[i+1:]:
                            new_con, added = df_add_edge(st.session_state.connections_df, node1, node2)
                            if added:
                                st.session_state.connections_df = new_con
                                connections_added += 1
                            else:
                                connections_already_exist += 1
                    
                    st.session_state.routes_df = None
                    
                    if connections_added > 0:
                        message = f"Added {connections_added} new connection{'s' if connections_added != 1 else ''}"
                        if connections_already_exist > 0:
                            message += f" ({connections_already_exist} already existed)"
                        st.success(message)
                        st.session_state.graph_key_v += 1
                        st.rerun()
                    elif connections_already_exist > 0:
                        st.info(f"All {connections_already_exist} possible connections already exist.")
                    else:
                        st.warning("Could not add any connections.")
                
                
                # Move Multiple Nodes
                st.markdown("**Move Multiple Nodes**")
                st.caption(f"{len(sel_nodes2)} nodes selected")
                
                col1, col2 = st.columns(2)
                with col1:
                    dx = st.number_input("Offset X", value=0.0, step=10.0, key="multi_offset_dx")
                with col2:
                    dy = -st.number_input("Offset Y", value=0.0, step=10.0, key="multi_offset_dy")
                
                if st.button("Apply offset to selected", key="apply_offset_btn", width="stretch"):
                    st.session_state.tray_df = apply_offset_to_nodes(
                        st.session_state.tray_df,
                        sel_nodes2,
                        float(dx),
                        float(dy)
                    )
                    st.session_state.layout_unsaved_hint = True
                    st.session_state.graph_key_v += 1
                    st.success(f"Offset applied! Moved {len(sel_nodes2)} nodes by ({dx}, {dy})")
                    st.rerun()
                
                # Scale Multiple Nodes
                st.markdown("**Scale Multiple Nodes**")
                st.caption("Scales nodes from their center point")
                
                # Initialize scale value in session state if not present
                if "scale_value" not in st.session_state:
                    st.session_state.scale_value = 1.0
                
                scale_col1, scale_col2 = st.columns([3, 1])
                with scale_col1:
                    st.session_state.scale_value = st.number_input(
                        "Scale factor",
                        value=st.session_state.scale_value,
                        min_value=0.1,
                        step=0.1,
                        key="scale_input"
                    )
                with scale_col2:
                    # Add vertical spacing to align button with number input
                    st.write("")
                    if st.button("Apply", key="apply_scale_btn", width="stretch"):
                        scale_factor = float(st.session_state.scale_value)
                        if scale_factor == 1.0:
                            pass
                        else:
                            st.session_state.tray_df = apply_scale_to_nodes(
                                st.session_state.tray_df,
                                sel_nodes2,
                                scale_factor
                            )
                            st.session_state.layout_unsaved_hint = True
                            st.session_state.graph_key_v += 1
                            st.success(f"Scale applied! Scaled {len(sel_nodes2)} nodes by factor {scale_factor:.2f}")
                            st.rerun()
                
                # Display warning if scale factor is 1.0
                if st.session_state.scale_value == 1.0:
                    st.info("Scale factor is 1.0 — no change will occur")
                
                # Get noise levels for all selected nodes
                noise_levels = []
                node_noise_map = {}
                for node_name in sel_nodes2:
                    node_name = str(node_name).strip()
                    tdf = st.session_state.tray_df
                    row = tdf[tdf["RunName"].astype(str).str.strip() == node_name]
                    if not row.empty:
                        nl = str(row.iloc[0].get("Noise Level", "")).strip()
                        noise_levels.append(nl)
                        node_noise_map[node_name] = nl
                
                # Always show noise level editing (regardless of whether levels match)
                st.markdown("**Edit Noise Level (All Selected)**")
                all_same = noise_levels and all(nl == noise_levels[0] for nl in noise_levels)
                if all_same:
                    st.caption("All selected nodes have the same noise level.")
                else:
                    st.caption("Selected nodes have different noise levels.")
                    with st.expander("Show current noise levels", expanded=False):
                        for node_name in sel_nodes2:
                            current_nl = node_noise_map.get(node_name, "(unknown)")
                            st.write(f"• {node_name}: {current_nl if current_nl else '(blank)'}")
                
                preset_map = {
                    "": "(leave as-is)",
                    "1": "1",
                    "2": "2",
                    "1,2": "1,2",
                    "2,1": "2,1",
                    "3": "3",
                    "4": "4",
                    "3,4": "3,4",
                    "4,3": "4,3",
                }
                presets = list(preset_map.keys())
                
                # Set default preset index based on first node's level
                current_text = str(noise_levels[0] or "").strip() if noise_levels else ""
                preset_index = 0
                if current_text in presets:
                    preset_index = presets.index(current_text)
                
                nl_preset_sel = st.selectbox(
                    "Preset",
                    options=presets,
                    index=preset_index,
                    key="multi_noise_preset",
                    format_func=lambda x: preset_map.get(x, x),
                )
                
                nl_custom_sel = st.text_input(
                    "Custom (optional: overrides preset if non-empty)",
                    value="",
                    key="multi_noise_custom",
                    placeholder="e.g. 1 or 2 or 1,2",
                )
                
                if st.button("Apply noise level to all", key="apply_multi_noise_btn", width="stretch"):
                    new_text = (nl_custom_sel or "").strip() if (nl_custom_sel or "").strip() else str(nl_preset_sel or "").strip()
                    if new_text == "(leave as-is)":
                        new_text = current_text
                    
                    success_count = 0
                    for node_name in sel_nodes2:
                        node_name = str(node_name).strip()
                        new_df, ok = df_set_node_noise_level(st.session_state.tray_df, node_name, new_text)
                        if ok:
                            st.session_state.tray_df = ensure_xy_columns(new_df)
                            success_count += 1
                    
                    if success_count > 0:
                        st.session_state.routes_df = None
                        st.success(f"Updated noise level for {success_count} nodes to: {new_text if new_text else '(blank)'}")
                        st.session_state.graph_key_v += 1
                        st.rerun()
                
                # Duplicate group
                if st.button("Duplicate selected group", key="dup_group_btn", width="stretch"):
                    new_df, new_names = df_duplicate_nodes(st.session_state.tray_df, sel_nodes2)
                    st.session_state.tray_df = ensure_xy_columns(new_df)
                    st.session_state.routes_df = None
                    st.success(f"Duplicated {len(sel_nodes2)} nodes. New names: {', '.join(new_names)}")
                    st.session_state.graph_key_v += 1
                    st.rerun()

            elif sel_nodes2:
                node_id = str(sel_nodes2[0]).strip()
                tdf = st.session_state.tray_df
                row = tdf[tdf["RunName"].astype(str).str.strip() == node_id]
                noise_val = ""
                x_val = y_val = None
                email_val = phone_val = role_val = notes_val = ""
                if not row.empty:
                    noise_val = row.iloc[0].get("Noise Level", "")
                    x_val = row.iloc[0].get("X", pd.NA)
                    y_val = row.iloc[0].get("Y", pd.NA)
                    email_val = str(row.iloc[0].get("Email", "") or "").strip()
                    phone_val = str(row.iloc[0].get("Phone", "") or "").strip()
                    role_val = str(row.iloc[0].get("Role", "") or "").strip()
                    notes_val = str(row.iloc[0].get("Notes", "") or "").strip()

                st.markdown("**Selected Node**")
                st.markdown("**Name:**")
                st.code(node_id, language=None)
                st.write("**Type:** Person")
                st.write(f"**Noise Level:** {noise_val if str(noise_val).strip() else 'N/A'}")
                st.write(f"**X:** {'' if pd.isna(x_val) else x_val}")
                st.write(f"**Y:** {'' if pd.isna(y_val) else y_val}")
                if role_val:
                    st.write(f"**Role:** {role_val}")
                if email_val:
                    st.write(f"**Email:** {email_val}")
                if phone_val:
                    st.write(f"**Phone:** {phone_val}")
                if notes_val:
                    st.write(f"**Notes:** {notes_val}")

                st.markdown("**Edit Person Details**")
                edit_email = st.text_input("Email", value=email_val, key="sel_edit_email")
                edit_phone = st.text_input("Phone", value=phone_val, key="sel_edit_phone")
                edit_role = st.text_input("Role", value=role_val, key="sel_edit_role")
                edit_notes = st.text_area("Notes", value=notes_val, key="sel_edit_notes")
                if st.button("Save person details", key="save_person_details_btn", width="stretch"):
                    node_mask = st.session_state.tray_df["RunName"].astype(str).str.strip() == node_id
                    st.session_state.tray_df.loc[node_mask, "Email"] = edit_email
                    st.session_state.tray_df.loc[node_mask, "Phone"] = edit_phone
                    st.session_state.tray_df.loc[node_mask, "Role"] = edit_role
                    st.session_state.tray_df.loc[node_mask, "Notes"] = edit_notes
                    st.success(f"Updated details for {node_id}")
                    st.session_state.graph_key_v += 1
                    st.rerun()

                eps = tray_to_endpoints.get(node_id, [])
                if eps:
                    st.caption("Endpoints connected to this tray/conduit (from Endpoints sheet):")
                    st.write(", ".join(eps))
                else:
                    st.caption("No endpoints reference this tray/conduit (from Endpoints sheet).")

                st.markdown("**Connect this node**")
                connect_options = [""] + [n for n in node_names if n != node_id]
                target = st.selectbox(
                    "Connect to (type to search)",
                    options=connect_options,
                    index=0,
                    key="sel_connect_target",
                )

                c_add, c_del = st.columns(2)
                with c_add:
                    if st.button("Add connection", key="sel_add_conn_btn", width="stretch"):
                        tgt = (target or "").strip()
                        if not tgt:
                            st.warning("Choose a target node.")
                        else:
                            new_con, added = df_add_edge(st.session_state.connections_df, node_id, tgt)
                            if added:
                                st.session_state.connections_df = new_con
                                st.session_state.routes_df = None
                                st.success(f"Added connection: {node_id} ↔ {tgt}")
                                st.session_state.graph_key_v += 1
                                st.rerun()
                            else:
                                st.info("That connection already exists (or could not be added).")
                with c_del:
                    if st.button("Delete connection", key="sel_del_conn_btn", width="stretch"):
                        tgt = (target or "").strip()
                        if not tgt:
                            st.warning("Choose a target node.")
                        else:
                            before = len(st.session_state.connections_df) if st.session_state.connections_df is not None else 0
                            st.session_state.connections_df = df_delete_edge(st.session_state.connections_df, node_id, tgt)
                            after = len(st.session_state.connections_df) if st.session_state.connections_df is not None else 0
                            st.session_state.routes_df = None
                            if before == after:
                                st.info("No matching connection was found to delete.")
                            else:
                                st.success(f"Deleted connection: {node_id} ↔ {tgt}")
                                st.session_state.graph_key_v += 1
                            st.rerun()

                if st.button("Delete node", key="sel_delete_node", width="stretch"):
                    t, c, e = df_delete_node(
                        st.session_state.tray_df,
                        st.session_state.connections_df,
                        st.session_state.endpoints_df,
                        node_id
                    )
                    st.session_state.tray_df = ensure_xy_columns(t)
                    st.session_state.connections_df = c
                    st.session_state.endpoints_df = e
                    st.session_state.routes_df = None
                    if st.session_state.focus_node == node_id:
                        st.session_state.focus_node = None
                    st.session_state.sel_nodes = []
                    st.session_state.sel_edges = []
                    st.success("Node deleted.")
                    st.session_state.graph_key_v += 1
                    st.rerun()

                st.markdown("**Rename**")
                new_name = st.text_input("New name", value=node_id, key="sel_rename_node_new")
                if st.button("Rename node", key="sel_rename_node_btn", width="stretch"):
                    if new_name.strip() and new_name.strip() != node_id:
                        if (st.session_state.tray_df["RunName"].astype(str).str.strip() == new_name.strip()).any():
                            st.error("That name already exists.")
                        else:
                            t, c, e = df_rename_node(
                                st.session_state.tray_df,
                                st.session_state.connections_df,
                                st.session_state.endpoints_df,
                                node_id,
                                new_name.strip()
                            )
                            st.session_state.tray_df = ensure_xy_columns(t)
                            st.session_state.connections_df = c
                            st.session_state.endpoints_df = e
                            st.session_state.routes_df = None
                            if st.session_state.focus_node == node_id:
                                st.session_state.focus_node = new_name.strip()
                            st.session_state.sel_nodes = [new_name.strip()]
                            st.session_state.sel_edges = []
                            st.success("Node renamed.")
                            st.session_state.graph_key_v += 1
                            st.rerun()
                    else:
                        st.warning("Enter a different non-empty name.")

                st.markdown("**Duplicate**")
                dup_name = st.text_input("Duplicate as", value=f"{node_id}_COPY", key="sel_dup_node_new")
                if st.button("Duplicate node", key="sel_dup_node_btn", width="stretch"):
                    if dup_name.strip():
                        if (st.session_state.tray_df["RunName"].astype(str).str.strip() == dup_name.strip()).any():
                            st.error("That duplicate name already exists.")
                        else:
                            st.session_state.tray_df = df_duplicate_node(st.session_state.tray_df, node_id, dup_name.strip())
                            st.session_state.routes_df = None
                            st.success("Node duplicated.")
                            st.session_state.graph_key_v += 1
                            st.rerun()
                    else:
                        st.warning("Enter a name for the duplicate.")

                st.markdown("**Edit noise level**")

                preset_map = {
                    "": "(leave as-is)",
                    "1": "1",
                    "2": "2",
                    "1,2": "1,2",
                    "2,1": "2,1",
                    "3": "3",
                    "4": "4",
                    "3,4": "3,4",
                    "4,3": "4,3",
                }
                presets = list(preset_map.keys())

                current_text = str(noise_val or "").strip()
                preset_index = 0
                if current_text in presets:
                    preset_index = presets.index(current_text)

                nl_preset_sel = st.selectbox(
                    "Preset",
                    options=presets,
                    index=preset_index,
                    key="sel_noise_preset",
                    format_func=lambda x: preset_map.get(x, x),
                )

                nl_custom_sel = st.text_input(
                    "Custom (optional: overrides preset if non-empty)",
                    value="",
                    key="sel_noise_custom",
                    placeholder="e.g. 1 or 2 or 1,2",
                )

                if st.button("Apply noise level", key="apply_noise_btn", width="stretch"):
                    new_text = (nl_custom_sel or "").strip() if (nl_custom_sel or "").strip() else str(nl_preset_sel or "").strip()
                    if new_text == "(leave as-is)":
                        new_text = current_text

                    new_df, ok = df_set_node_noise_level(st.session_state.tray_df, node_id, new_text)
                    if ok:
                        st.session_state.tray_df = ensure_xy_columns(new_df)
                        st.session_state.routes_df = None
                        st.success(f"Updated noise level for {node_id} to: {new_text if new_text else '(blank)'}")
                        st.session_state.graph_key_v += 1
                        st.rerun()
                    else:
                        st.error("Could not update noise level for the selected node.")

            elif sel_edges2:
                raw_edge = sel_edges2[0]
                a, b, disp = parse_selected_edge(raw_edge)

                st.markdown("**Selected Connection**")
                st.write(f"**Connection:** {disp}")

                if not (a and b):
                    st.caption("Could not parse endpoints for this connection (unexpected format).")
                    st.code(str(raw_edge))
                else:
                    st.write(f"**From:** {a}")
                    st.write(f"**To:** {b}")
                    st.divider()
                    
                    # Exposed checkbox
                    current_exposed = get_edge_exposed_status(st.session_state.connections_df, a, b)
                    new_exposed = st.checkbox(
                        "Exposed Conduit Route",
                        value=current_exposed,
                        key="edge_exposed_checkbox",
                    )
                    
                    if new_exposed != current_exposed:
                        st.session_state.connections_df = set_edge_exposed_status(
                            st.session_state.connections_df, a, b, new_exposed
                        )
                        st.session_state.routes_df = None
                        st.session_state.graph_key_v += 1
                        st.success("Exposed status updated. Rerendering graph...")
                        st.rerun()
                    
                    if st.button("Delete connection", key="sel_delete_edge", width="stretch"):
                        before = len(st.session_state.connections_df) if st.session_state.connections_df is not None else 0
                        st.session_state.connections_df = df_delete_edge(st.session_state.connections_df, a, b)
                        after = len(st.session_state.connections_df) if st.session_state.connections_df is not None else 0
                        st.session_state.routes_df = None

                        st.session_state.sel_nodes = []
                        st.session_state.sel_edges = []

                        if before == after:
                            st.info("No matching connection was found to delete (already removed?).")
                        else:
                            st.success("Connection deleted.")
                        st.session_state.graph_key_v += 1
                        st.rerun()

                    st.caption("This deletes the row from the Connections sheet (undirected match).")

            else:
                st.info("Select a node or connection in the graph to edit it here.")

            with st.expander("🔍 Search / Focus", expanded=False):
                st.session_state.focus_depth = st.slider(
                    "Neighborhood depth",
                    min_value=1,
                    max_value=6,
                    value=int(st.session_state.focus_depth),
                    step=1,
                )

                focus_pick = st.selectbox(
                    "Find a node (type to search)",
                    options=[""] + node_names,
                    index=0 if not st.session_state.focus_node else (
                        node_names.index(st.session_state.focus_node) + 1
                        if st.session_state.focus_node in node_names else 0
                    ),
                    key="focus_pick_select",
                )

                c1, c2 = st.columns(2)
                with c1:
                    if st.button("Focus", width="stretch", key="focus_btn"):
                        if focus_pick and focus_pick.strip():
                            st.session_state.focus_node = focus_pick.strip()
                            st.session_state.sel_nodes = [st.session_state.focus_node]
                            st.session_state.sel_edges = []
                            st.session_state.graph_key_v += 1
                            st.rerun()
                        else:
                            st.warning("Pick a node to focus.")
                with c2:
                    if st.button("Clear", width="stretch", key="clear_focus_btn"):
                        st.session_state.focus_node = None
                        st.session_state.graph_key_v += 1
                        st.rerun()

                if st.session_state.focus_node:
                    st.caption(f"Focused on: **{st.session_state.focus_node}** (showing local neighborhood)")

            with st.expander("🔗 Connection", expanded=False):
                conn_from = st.selectbox("From (type to search)", options=[""] + node_names, index=0, key="conn_from")
                conn_to = st.selectbox("To (type to search)", options=[""] + node_names, index=0, key="conn_to")

                b_add, b_del = st.columns(2)
                with b_add:
                    if st.button("Add", width="stretch", key="add_edge_btn"):
                        a = (conn_from or "").strip()
                        b = (conn_to or "").strip()
                        if not a or not b:
                            st.warning("Choose both From and To.")
                        elif a == b:
                            st.warning("From and To must be different.")
                        else:
                            new_con, added = df_add_edge(st.session_state.connections_df, a, b)
                            if added:
                                st.session_state.connections_df = new_con
                                st.session_state.routes_df = None
                                st.success(f"Added connection: {a} ↔ {b}")
                                st.session_state.graph_key_v += 1
                                st.rerun()
                            else:
                                st.info("That connection already exists (or could not be added).")
                with b_del:
                    if st.button("Delete", width="stretch", key="delete_edge_btn"):
                        a = (conn_from or "").strip()
                        b = (conn_to or "").strip()
                        if not a or not b:
                            st.warning("Choose both From and To.")
                        elif a == b:
                            st.warning("From and To must be different.")
                        else:
                            before = len(st.session_state.connections_df) if st.session_state.connections_df is not None else 0
                            st.session_state.connections_df = df_delete_edge(st.session_state.connections_df, a, b)
                            after = len(st.session_state.connections_df) if st.session_state.connections_df is not None else 0
                            st.session_state.routes_df = None
                            if before == after:
                                st.info("No matching connection was found to delete.")
                            else:
                                st.session_state.sel_edges = []
                                st.session_state.sel_nodes = []
                                st.success(f"Deleted connection: {a} ↔ {b}")
                                st.session_state.graph_key_v += 1
                            st.rerun()

            with st.expander("➕ Add node", expanded=False):
                new_node_name = st.text_input("New node name", value="", key="add_node_name")
                nl_preset = st.selectbox("Noise level (preset)", options=["(leave blank)", "1", "2", "1,2", "3", "4", "3,4"], index=0, key="add_node_nl_preset")
                nl_custom = st.text_input("Noise level (custom text, optional)", value="", key="add_node_nl_custom")

                p1, p2 = st.columns(2)
                with p1:
                    new_x = st.text_input("X (optional)", value="", key="add_node_x")
                with p2:
                    new_y = st.text_input("Y (optional)", value="", key="add_node_y")

                if st.button("Add node", width="stretch", key="add_node_btn"):
                    name = (new_node_name or "").strip()
                    if not name:
                        st.warning("Enter a node name.")
                    else:
                        if (nl_custom or "").strip():
                            nl_text = nl_custom.strip()
                        else:
                            nl_text = "" if nl_preset == "(leave blank)" else nl_preset

                        new_df, added = df_add_node(
                            st.session_state.tray_df,
                            name=name,
                            noise_level_text=nl_text,
                            x=new_x,
                            y=new_y,
                        )

                        if not added:
                            st.error("Could not add node (it may already exist).")
                        else:
                            st.session_state.tray_df = ensure_xy_columns(new_df)
                            st.session_state.routes_df = None
                            st.session_state.focus_node = None
                            st.session_state.sel_nodes = [name]
                            st.session_state.sel_edges = []
                            st.success(f"Added node: {name}")
                            st.session_state.graph_key_v += 1
                            st.rerun()

            with st.expander("🗑️ Delete node", expanded=False):
                del_pick = st.selectbox("Node to delete (type to search)", options=[""] + node_names, index=0, key="delete_node_pick")
                if st.button("Delete selected node", width="stretch", key="delete_node_autofill_btn"):
                    if not del_pick.strip():
                        st.warning("Choose a node to delete.")
                    else:
                        node_id = del_pick.strip()
                        t, c, e = df_delete_node(
                            st.session_state.tray_df,
                            st.session_state.connections_df,
                            st.session_state.endpoints_df,
                            node_id
                        )
                        st.session_state.tray_df = ensure_xy_columns(t)
                        st.session_state.connections_df = c
                        st.session_state.endpoints_df = e
                        st.session_state.routes_df = None
                        if st.session_state.focus_node == node_id:
                            st.session_state.focus_node = None
                        if st.session_state.sel_nodes and st.session_state.sel_nodes[0] == node_id:
                            st.session_state.sel_nodes = []
                            st.session_state.sel_edges = []
                        st.success(f"Deleted node: {node_id}")
                        st.session_state.graph_key_v += 1
                        st.rerun()

            with st.expander("📍 Endpoint lookup", expanded=False):
                ep_pick = st.selectbox(
                    "Choose an endpoint/device (type to search)",
                    options=[""] + endpoint_names,
                    index=0,
                    key="endpoint_lookup_pick",
                )

                e1, e2 = st.columns([1, 1])
                with e1:
                    if st.button("Highlight connected trays/conduits", width="stretch", key="endpoint_lookup_highlight_btn"):
                        if not (ep_pick or "").strip():
                            st.warning("Pick an endpoint/device.")
                        else:
                            ep = (ep_pick or "").strip().lstrip("+")
                            trays = endpoint_to_trays.get(ep, [])
                            trays_present = [t for t in trays if t in node_names]
                            st.session_state.endpoint_highlight_nodes = set(trays_present)
                            st.session_state.endpoint_highlight_ep = ep
                            st.session_state.graph_key_v += 1
                            if trays_present:
                                st.success(f"Highlighted {len(trays_present)} tray/conduit node(s) for {ep}.")
                            else:
                                st.info("No trays/conduits for that endpoint are present in Tray.RunName (nothing to highlight).")
                            st.rerun()

                with e2:
                    if st.button("Clear endpoint highlight", width="stretch", key="endpoint_lookup_clear_highlight_btn"):
                        st.session_state.endpoint_highlight_nodes = set()
                        st.session_state.endpoint_highlight_ep = None
                        st.session_state.graph_key_v += 1
                        st.success("Cleared endpoint highlight.")
                        st.rerun()

                if st.session_state.endpoint_highlight_ep:
                    st.caption(f"Endpoint highlight active for: **{st.session_state.endpoint_highlight_ep}**")
                    if st.session_state.endpoint_highlight_nodes:
                        st.write(", ".join(sorted(st.session_state.endpoint_highlight_nodes)))
                    else:
                        st.write("No nodes are currently highlighted for this endpoint.")



            with st.expander("🧭 Generate route", expanded=False):
                st.session_state.gen_route_mode = st.radio(
                    "Mode",
                    options=["Auto-route (endpoints)", "Manual (click nodes in graph)"],
                    index=0 if st.session_state.gen_route_mode == "Auto-route (endpoints)" else 1,
                    key="gen_route_mode_radio",
                )

                st.session_state.gen_route_noise_level = st.selectbox(
                    "Noise Level",
                    options=[1, 2, 3, 4],
                    index=[1, 2, 3, 4].index(int(st.session_state.gen_route_noise_level) if st.session_state.gen_route_noise_level else 1),
                    key="gen_route_noise_level_pick",
                )

                if st.session_state.gen_route_mode == "Auto-route (endpoints)":
                    st.caption("Two search bars: choose Endpoint OR Tray/Conduit on each side.")

                    start_pick = st.selectbox(
                        "Start (endpoint OR tray/conduit)",
                        options=combined_options,
                        index=0,
                        key="gen_route_auto_start_pick_select",
                        format_func=_format_autoroute_option,
                    )

                    end_pick = st.selectbox(
                        "End (endpoint OR tray/conduit)",
                        options=combined_options,
                        index=0,
                        key="gen_route_auto_end_pick_select",
                        format_func=_format_autoroute_option,
                    )

                    g1, g2 = st.columns([1, 1])
                    with g1:
                        if st.button("Generate auto-route", width="stretch", key="gen_route_auto_btn"):
                            nl = int(st.session_state.gen_route_noise_level)

                            path, msg = try_auto_route_packed(
                                st.session_state.tray_df,
                                st.session_state.connections_df,
                                st.session_state.endpoints_df,
                                start_pick,
                                end_pick,
                                nl,
                            )

                            st.session_state.gen_route_auto_msg = msg
                            if path:
                                present = [n for n in path if n in node_names]
                                st.session_state.gen_route_auto_nodes = set(present)
                                st.session_state.gen_route_auto_via = format_route_nodes_for_via(path, noise_level=nl)
                                st.session_state.graph_key_v += 1
                                st.success("Auto-route generated and highlighted.")
                                st.rerun()
                            else:
                                st.session_state.gen_route_auto_nodes = set()
                                st.session_state.gen_route_auto_via = ""
                                st.session_state.graph_key_v += 1
                                st.warning(msg)
                                st.rerun()

                    with g2:
                        if st.button("Clear generated auto-route", width="stretch", key="gen_route_auto_clear_btn"):
                            st.session_state.gen_route_auto_nodes = set()
                            st.session_state.gen_route_auto_via = ""
                            st.session_state.gen_route_auto_msg = ""
                            st.session_state.graph_key_v += 1
                            st.success("Cleared generated auto-route.")
                            st.rerun()

                    if st.session_state.gen_route_auto_msg and not st.session_state.gen_route_auto_via:
                        st.markdown("**Why it failed**")
                        st.code(st.session_state.gen_route_auto_msg)

                    if st.session_state.gen_route_auto_via:
                        st.markdown("**Generated Via**")
                        st.code(st.session_state.gen_route_auto_via)

                else:
                    st.caption(
                        "Click nodes in the graph (one-by-one) to build a route. "
                        "Nodes will highlight as you click, and the Via string updates live."
                    )

                    m1, m2 = st.columns([1, 1])
                    with m1:
                        if st.button("Clear manual route", width="stretch", key="gen_route_manual_clear_btn"):
                            st.session_state.gen_route_manual_nodes = []
                            st.session_state.gen_route_manual_via = ""
                            st.session_state.gen_route_manual_last_clicked = None
                            st.session_state.graph_key_v += 1
                            st.success("Cleared manual route.")
                            st.rerun()

                    with m2:
                        if st.button("Clear ALL generated routes", width="stretch", key="gen_route_clear_all_btn"):
                            st.session_state.gen_route_auto_nodes = set()
                            st.session_state.gen_route_auto_via = ""
                            st.session_state.gen_route_auto_msg = ""
                            st.session_state.gen_route_manual_nodes = []
                            st.session_state.gen_route_manual_via = ""
                            st.session_state.gen_route_manual_last_clicked = None
                            st.session_state.graph_key_v += 1
                            st.success("Cleared auto + manual generated routes.")
                            st.rerun()

                    if st.session_state.gen_route_manual_nodes:
                        ok, bad_pairs = validate_manual_route_steps(st.session_state.connections_df, st.session_state.gen_route_manual_nodes)
                        if not ok:
                            st.warning(
                                "Manual route has step(s) that are not directly connected in Connections:\n\n"
                                + "\n".join([f"- {a} ↔ {b}" for a, b in bad_pairs])
                            )

                    st.markdown("**Manual route nodes**")
                    st.write(" → ".join(st.session_state.gen_route_manual_nodes) if st.session_state.gen_route_manual_nodes else "(none yet)")

                    st.markdown("**Generated Via (live)**")
                    st.code(st.session_state.gen_route_manual_via or "")


            with st.expander("📐 Layout", expanded=False):
                if st.button("Home / Recenter graph view", key="home_recenter_btn", width="stretch"):
                    st.session_state.focus_node = None
                    st.session_state.sel_nodes = []
                    st.session_state.sel_edges = []
                    st.session_state.graph_key_v += 1
                    st.success("Recentered graph view.")
                    st.rerun()

                if not st.session_state.layout_opt_active:
                    if st.button("Optimize layout (turns on physics)", key="opt_turn_on_physics_btn", width="stretch"):
                        tdf = ensure_xy_columns(st.session_state.tray_df).copy()
                        backup = {
                            "RunName": tdf["RunName"].astype(str).tolist(),
                            "X": tdf["X"].tolist(),
                            "Y": tdf["Y"].tolist(),
                        }
                        st.session_state.layout_opt_backup_xy = backup

                        st.session_state.tray_df["X"] = pd.NA
                        st.session_state.tray_df["Y"] = pd.NA

                        st.session_state.initial_layout_autosave_active = False

                        st.session_state.routes_df = None
                        st.session_state.focus_node = None
                        st.session_state.sel_nodes = []
                        st.session_state.sel_edges = []

                        st.session_state.layout_opt_active = True
                        st.session_state.layout_opt_last_positions = None
                        st.session_state.layout_unsaved_hint = False
                        st.session_state.graph_key_v += 1
                        st.success("Physics enabled for optimization. Let it settle, then save or cancel.")
                        st.rerun()
                else:
                    st.info(
                        "Optimization is active (physics ON).\n\n"
                        "✅ Let the graph settle, then click **Save optimized positions (turns off physics)**.\n"
                        "↩️ Or click **Cancel optimization (turns off physics)** to revert to the previous layout."
                    )

                    if st.button("Save optimized positions (turns off physics)", key="save_opt_positions_btn", width="stretch"):
                        positions2 = st.session_state.layout_opt_last_positions

                        if positions2 and isinstance(positions2, dict) and len(positions2) > 0:
                            st.session_state.tray_df = apply_positions_to_tray(st.session_state.tray_df, positions2)
                            st.session_state.layout_opt_active = False
                            st.session_state.layout_opt_last_positions = None
                            st.session_state.layout_opt_backup_xy = None
                            st.session_state.layout_unsaved_hint = False
                            st.session_state.graph_key_v += 1
                            st.success(f"Saved optimized positions for {len(positions2)} node(s). Physics is now OFF.")
                            st.rerun()
                        else:
                            st.warning(
                                "No positions were returned yet.\n\n"
                                "✅ Wait a moment for stabilization, then click **Save optimized positions** again."
                            )

                    if st.button("Cancel optimization (turns off physics)", key="cancel_opt_btn", width="stretch"):
                        backup = st.session_state.layout_opt_backup_xy
                        if backup and "RunName" in backup:
                            tdf = ensure_xy_columns(st.session_state.tray_df).copy()
                            rn = tdf["RunName"].astype(str).tolist()
                            bmap = {str(n).strip(): (backup["X"][i], backup["Y"][i]) for i, n in enumerate(backup["RunName"])}

                            xs = []
                            ys = []
                            for n in rn:
                                x, y = bmap.get(str(n).strip(), (pd.NA, pd.NA))
                                xs.append(x)
                                ys.append(y)
                            tdf["X"] = xs
                            tdf["Y"] = ys
                            st.session_state.tray_df = tdf

                        st.session_state.layout_opt_active = False
                        st.session_state.layout_opt_last_positions = None
                        st.session_state.layout_opt_backup_xy = None
                        st.session_state.layout_unsaved_hint = False
                        st.session_state.graph_key_v += 1
                        st.success("Optimization canceled. Previous layout restored. Physics is now OFF.")
                        st.rerun()

            # ------------------------------------------------------------
            # Save current positions
            # FIXED: after successful save, force graph re-mount so it
            # shows the saved layout immediately (no “stale view”).
            # Also reset graph_interaction_sig to avoid auto-clearing
            # the success message on that forced refresh.
            # (Note: save_clicked and save_notice_slot are now defined at top of tools_box)
            # Save button click handler (defined at top of tools_box, logic here)
            if save_clicked:
                pos = None
                sn = se = None
                try:
                    if selection:
                        sn, se, pos = selection
                except Exception:
                    sn, se, pos = None, None, None

                if isinstance(pos, dict) and len(pos) > 0:
                    st.session_state.tray_df = apply_positions_to_tray(st.session_state.tray_df, pos)
                    st.session_state.layout_unsaved_hint = False

                    # Keep message visible + immediately refresh graph instance
                    st.session_state.save_positions_notice = ("success", "✅ Positions saved to Tray (X/Y).")

                    # IMPORTANT: prevent the forced refresh from being interpreted as "user interaction"
                    st.session_state.graph_interaction_sig = None

                    # IMPORTANT: remount the vis network so it uses saved X/Y right away
                    st.session_state.graph_key_v += 1

                    # optional: clear selection so user sees a clean graph post-save
                    st.session_state.sel_nodes = []
                    st.session_state.sel_edges = []

                    st.rerun()
                else:
                    sel_active = bool(
                        (sn and len(sn) > 0) or
                        (se and len(se) > 0) or
                        (st.session_state.sel_nodes) or
                        (st.session_state.sel_edges)
                    )
                    if sel_active:
                        st.session_state.save_positions_notice = (
                            "warning",
                            "No positions were returned.\n\n"
                            "✅ Fix:\n"
                            "- Click **empty space in the graph** (to deselect the node/connection)\n"
                            "- Then click **Save** again."
                        )
                    else:
                        st.session_state.save_positions_notice = (
                            "warning",
                            "No positions were returned yet.\n\n"
                            "✅ Try dragging any node slightly, then click **Save** again."
                        )

            if st.session_state.save_positions_notice:
                lvl, txt = st.session_state.save_positions_notice
                if lvl == "success":
                    save_notice_slot.success(txt)
                elif lvl == "warning":
                    save_notice_slot.warning(txt)
                elif lvl == "info":
                    save_notice_slot.info(txt)
                else:
                    save_notice_slot.error(txt)

    st.divider()
    st.caption(
        "If the workbook loads with blank X/Y, the app will automatically lock in the FIRST layout once. "
        "After that, positions only change when you click **Save current node positions to Tray (X/Y)**."
    )
    st.caption(
        "**Multi-select:** Hold **Shift** and left-click drag to select multiple nodes together."
    )

with tab5:
    st.subheader("Route cables")
    colA, colB = st.columns([1, 1])

    with colA:
        if st.button("Route Now", key="route_btn", use_container_width=True):
            try:
                st.session_state.routes_df = compute_routes_df(
                    st.session_state.tray_df,
                    st.session_state.connections_df,
                    st.session_state.endpoints_df,
                    st.session_state.cables_df,
                )
                st.success("Routing complete.")
            except Exception as e:
                st.error(f"Routing failed: {e}")

    routes_df = st.session_state.routes_df
    if routes_df is not None:
        st.dataframe(routes_df, width="stretch")

        dfs_for_export = {
            "Tray": st.session_state.tray_df,
            "Connections": st.session_state.connections_df,
            "Endpoints": st.session_state.endpoints_df,
            "Cables(input)": st.session_state.cables_df,
        }

        routed_bytes = write_routed_workbook_bytes(dfs_for_export, routes_df)

        st.download_button(
            "Download ROUTED workbook (includes CableRoutes(output))",
            data=routed_bytes,
            file_name="network_configuration_routed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.info("Click **Route Now** to generate CableRoutes(output).")

with tab_reverse:
    st.subheader("Reverse Engineer from Routed Workbook")
    st.write(
        "Upload a routed Excel workbook (with a 'CableRoutes(output)' sheet) to reverse-engineer "
        "and reconstruct the Tray, Connections, and Endpoints sheets."
    )
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        re_uploaded = st.file_uploader(
            "Upload routed workbook",
            type="xlsx",
            key="reverse_engineer_uploader"
        )
    
    with col2:
        st.write("")  # Align with uploader height
        reverse_engineer_btn = st.button("Reverse Engineer", key="reverse_engineer_btn")
    
    if reverse_engineer_btn and re_uploaded is not None:
        try:
            re_file_bytes = re_uploaded.getvalue()
            from io import BytesIO
            
            (
                re_tray_df,
                re_connections_df,
                re_endpoints_df,
                re_node_levels,
                re_connections,
                re_cables_original,
                re_routes_original,
            ) = reverse_engineer_from_routes(BytesIO(re_file_bytes))
            
            st.success("Reverse engineering complete!")
            
            # Display results
            st.subheader("Reconstructed Tray")
            st.dataframe(re_tray_df, width='stretch')
            
            st.subheader("Reconstructed Connections")
            st.dataframe(re_connections_df, width='stretch')
            
            st.subheader("Reconstructed Endpoints")
            st.dataframe(re_endpoints_df, width='stretch')
            
            # Create download for reconstructed workbook
            re_output = BytesIO()
            with pd.ExcelWriter(re_output, engine="openpyxl") as writer:
                re_tray_df.to_excel(writer, sheet_name="Tray", index=False)
                re_connections_df.to_excel(writer, sheet_name="Connections", index=False)
                re_endpoints_df.to_excel(writer, sheet_name="Endpoints", index=False)
                
                # Add empty Cables(input) sheet if not present
                if re_cables_original is not None:
                    re_cables_original.to_excel(writer, sheet_name="Cables(input)", index=False)
                else:
                    empty_cables = pd.DataFrame(columns=["Cable number", "equipfrom", "equipto", "Noise Level", "Sort", "INCLUDE", "EXCLUDE"])
                    empty_cables.to_excel(writer, sheet_name="Cables(input)", index=False)
                
                # Include original CableRoutes(output) for reference
                if re_routes_original is not None:
                    re_routes_original.to_excel(writer, sheet_name="CableRoutes(output)", index=False)
            
            re_output.seek(0)
            st.download_button(
                "Download Reconstructed Workbook",
                data=re_output.getvalue(),
                file_name="reconstructed_network.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_reconstructed_btn"
            )
            
        except Exception as e:
            st.error(f"Reverse engineering failed: {e}")
            import traceback
            st.write(traceback.format_exc())
    elif reverse_engineer_btn and re_uploaded is None:
        st.warning("Please upload a file first.")
